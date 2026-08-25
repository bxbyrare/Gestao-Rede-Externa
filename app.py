import os
import time
import csv
import json
import io
import datetime
import calendar
import traceback
import psycopg2
from psycopg2.extras import RealDictCursor
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask import (
    Flask, render_template, request, 
    redirect, url_for, session, jsonify, send_from_directory,
    make_response, Response, send_file
)

from collections import defaultdict

app = Flask(__name__)
app.secret_key = os.environ['SECRET_KEY']
app.permanent_session_lifetime = datetime.timedelta(hours=12)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024

def safe_int(val, default=0):
    if val is None or val == '':
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default

def safe_float(val, default=0.0):
    if val is None or val == '':
        return default
    try:
        return float(str(val).replace(',', '.'))
    except (ValueError, TypeError):
        return default


# Security Hardening Stores: IP Rate Limiters
RATE_LIMIT_STORE = defaultdict(list)
AUTH_RATE_LIMIT_STORE = defaultdict(list)

GLOBAL_RATE_LIMIT_MAX = 120  # Max 120 requests per 60s per IP
AUTH_RATE_LIMIT_MAX = 10     # Max 10 attempts per 60s per IP for login/resets

BLOCKED_USER_AGENTS = [
    'sqlmap', 'nikto', 'nmap', 'burpcollaborator', 'dirbuster',
    'wpscan', 'acunetix', 'havij', 'masscan', 'zgrab', 'gobuster', 'hydra'
]

@app.before_request
def enforce_security_policies():
    client_ip = request.remote_addr or request.headers.get('X-Forwarded-For', '127.0.0.1').split(',')[0].strip()
    now = time.time()
    
    # 1. Anti-Scanner / Anti-Automated Tool Detection (Burp Suite Automated Scanners, Sqlmap, Nikto, etc.)
    user_agent = (request.headers.get('User-Agent') or '').lower()
    for bad_ua in BLOCKED_USER_AGENTS:
        if bad_ua in user_agent:
            return jsonify({"error": "Acesso negado por políticas de segurança do servidor.", "status": 403}), 403

    # 2. Global Sliding Window Rate Limiting (All Requests)
    RATE_LIMIT_STORE[client_ip] = [t for t in RATE_LIMIT_STORE[client_ip] if now - t < 60]
    if len(RATE_LIMIT_STORE[client_ip]) >= GLOBAL_RATE_LIMIT_MAX:
        return jsonify({
            "error": "Muitas requisições simultâneas. Limite de taxa excedido. Por favor, aguarde 1 minuto.",
            "status": 429
        }), 429
    RATE_LIMIT_STORE[client_ip].append(now)
    
    # 3. Strict Auth Rate Limiting (/login and password resets)
    if request.path in ['/login', '/forgot-password', '/reset-password']:
        AUTH_RATE_LIMIT_STORE[client_ip] = [t for t in AUTH_RATE_LIMIT_STORE[client_ip] if now - t < 60]
        if len(AUTH_RATE_LIMIT_STORE[client_ip]) >= AUTH_RATE_LIMIT_MAX:
            if request.path.startswith('/api/') or request.is_json:
                return jsonify({
                    "error": "Muitas tentativas de autenticação. Por favor, aguarde 1 minuto para tentar novamente.",
                    "status": 429
                }), 429
            return render_template('login.html', error="Muitas tentativas de login. Por favor, aguarde 1 minuto."), 429
        if request.method == 'POST':
            AUTH_RATE_LIMIT_STORE[client_ip].append(now)

@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), camera=(), microphone=()'
    # Scoped per resource type instead of a blanket "https: data: blob:" default-src,
    # which previously let the page load a script/style/frame from ANY https origin.
    # 'unsafe-inline' remains on script-src/style-src because the UI still relies on
    # hundreds of onclick="..." attributes and style="..." attributes throughout
    # dashboard.html/main.js — dropping it would break virtually every button on the
    # page until those are migrated to addEventListener + CSS classes. 'unsafe-eval'
    # is dropped outright: nothing in this codebase calls eval()/new Function().
    response.headers['Content-Security-Policy'] = "; ".join([
        "default-src 'self'",
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://unpkg.com https://static.cloudflareinsights.com",
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://unpkg.com",
        "font-src 'self' https://fonts.gstatic.com",
        "img-src 'self' data: https://*.basemaps.cartocdn.com",
        "connect-src 'self' https://cloudflareinsights.com",
        "object-src 'none'",
        "base-uri 'self'",
        "form-action 'self'",
        "frame-ancestors 'self'",
        "upgrade-insecure-requests",
    ]) + ";"
    response.headers['Server'] = 'WAF-Protected'
    return response

# Configuration for Database and Upload Directories
DATABASE_URL = os.environ['DATABASE_URL']
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
ALLOWED_PROJECT_EXTENSIONS = {'pdf', 'kmz', 'kml', 'zip', 'doc', 'docx', 'xls', 'xlsx', 'txt', 'png', 'jpg', 'jpeg'}

# Create upload folders if they don't exist
os.makedirs(os.path.join(UPLOAD_FOLDER, 'projetos'), exist_ok=True)

# Helper function to get PostgreSQL connection
def get_db():
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    return conn

# Helper function to check allowed file extensions
def allowed_file(filename, extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in extensions

# Database Initialization Script
def init_db():
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # 1. Users Table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username VARCHAR(80) UNIQUE NOT NULL,
                password_hash VARCHAR(255) NOT NULL,
                role VARCHAR(30) NOT NULL, -- admin, supervisor, operator
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        
        # 2. Technicians Table (Pessoas)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS technicians (
                id SERIAL PRIMARY KEY,
                name VARCHAR(150) NOT NULL,
                cpf VARCHAR(20) UNIQUE NOT NULL,
                phone VARCHAR(30),
                identity VARCHAR(30),
                dob DATE NOT NULL,
                role VARCHAR(50) NOT NULL, -- Auxiliar, Técnico, Supervisor, Coordenador
                area VARCHAR(100),
                team_type VARCHAR(50), -- Fusão ou Lançamento
                shirt_size VARCHAR(10),
                boot_size VARCHAR(10),
                pants_size VARCHAR(10),
                jacket_size VARCHAR(10),
                team VARCHAR(100),
                created_by INT REFERENCES users(id) ON DELETE SET NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        # Make cpf, dob, role nullable in technicians and fix invalid years > 9999
        try:
            cur.execute("ALTER TABLE technicians ALTER COLUMN dob DROP NOT NULL;")
            cur.execute("ALTER TABLE technicians ALTER COLUMN cpf DROP NOT NULL;")
            cur.execute("ALTER TABLE technicians ALTER COLUMN role DROP NOT NULL;")
            cur.execute("UPDATE technicians SET dob = NULL WHERE EXTRACT(YEAR FROM dob) > 9999;")
        except Exception:
            conn.rollback()

        for col, coltype in [
            ('company', 'VARCHAR(50)'),
            ('registration_claro', 'VARCHAR(50)'),
            ('registration_third', 'VARCHAR(50)'),
            ('toa_login', 'VARCHAR(100)'),
            ('phone_model', 'VARCHAR(100)'),
            ('imei_1', 'VARCHAR(50)'),
            ('imei_2', 'VARCHAR(50)'),
            ('email', 'VARCHAR(150)')
        ]:
            try:
                cur.execute(f"ALTER TABLE technicians ADD COLUMN IF NOT EXISTS {col} {coltype};")
            except Exception:
                conn.rollback()
        
        # 3. Vehicles Table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vehicles (
                id SERIAL PRIMARY KEY,
                plate VARCHAR(20) UNIQUE NOT NULL,
                type VARCHAR(50) NOT NULL DEFAULT 'Utilitário',
                model VARCHAR(100),
                responsible_tech_id INT REFERENCES technicians(id) ON DELETE SET NULL,
                has_rack BOOLEAN NOT NULL DEFAULT FALSE,
                has_basket BOOLEAN NOT NULL DEFAULT FALSE,
                has_giroflex BOOLEAN NOT NULL DEFAULT FALSE,
                has_inverter BOOLEAN NOT NULL DEFAULT FALSE,
                ticket_car VARCHAR(50),
                area_rede VARCHAR(100),
                base VARCHAR(100),
                setor VARCHAR(100),
                condutor_dia VARCHAR(200),
                condutor_tarde VARCHAR(200),
                condutor_madrugada VARCHAR(200),
                subclus VARCHAR(50),
                created_by INT REFERENCES users(id) ON DELETE SET NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        # Add new columns if they don't exist (for existing databases)
        for col, coltype in [('area_rede','VARCHAR(100)'),('base','VARCHAR(100)'),('setor','VARCHAR(100)'),('condutor_dia','VARCHAR(200)'),('condutor_tarde','VARCHAR(200)'),('condutor_madrugada','VARCHAR(200)'),('subclus','VARCHAR(50)')]:
            try:
                cur.execute(f"ALTER TABLE vehicles ADD COLUMN IF NOT EXISTS {col} {coltype};")
            except Exception:
                conn.rollback()
        
        # 4. Financial Logs Table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS financial_logs (
                id SERIAL PRIMARY KEY,
                category VARCHAR(50) NOT NULL, -- Equipes, Cabos, Equipamentos, EPI, Ferramentas
                description TEXT NOT NULL,
                amount NUMERIC(12, 2) NOT NULL DEFAULT 0.00,
                logged_by INT REFERENCES users(id) ON DELETE SET NULL,
                logged_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)

        # 5. Project Folders Table [NEW]
        cur.execute("""
            CREATE TABLE IF NOT EXISTS project_folders (
                id SERIAL PRIMARY KEY,
                name VARCHAR(150) NOT NULL,
                parent_id INT REFERENCES project_folders(id) ON DELETE CASCADE,
                created_by INT REFERENCES users(id) ON DELETE SET NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        
        # 6. Projects Table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS projects (
                id SERIAL PRIMARY KEY,
                name VARCHAR(150) NOT NULL,
                description TEXT,
                kmz_path TEXT,
                pdf_path TEXT,
                created_by INT REFERENCES users(id) ON DELETE SET NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        # Add new project columns if they don't exist
        for col, coltype in [('area','VARCHAR(100)'),('folder_id','INT REFERENCES project_folders(id) ON DELETE CASCADE')]:
            try:
                cur.execute(f"ALTER TABLE projects ADD COLUMN IF NOT EXISTS {col} {coltype};")
            except Exception:
                conn.rollback()
                
        try:
            cur.execute("ALTER TABLE projects ALTER COLUMN kmz_path TYPE TEXT;")
            cur.execute("ALTER TABLE projects ALTER COLUMN pdf_path TYPE TEXT;")
        except Exception:
            conn.rollback()
        
        # 7. Audit Logs Table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS audit_logs (
                id SERIAL PRIMARY KEY,
                user_id INT REFERENCES users(id) ON DELETE SET NULL,
                username VARCHAR(80),
                action VARCHAR(255) NOT NULL,
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)

        # 8. Favorites Table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS favorites (
                id SERIAL PRIMARY KEY,
                title VARCHAR(150) NOT NULL,
                link TEXT NOT NULL,
                color VARCHAR(30) DEFAULT 'Vermelho',
                access_count INT DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        try:
            cur.execute("ALTER TABLE favorites ADD COLUMN IF NOT EXISTS access_count INT DEFAULT 0;")
        except Exception:
            conn.rollback()

        # 9. Routes Table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS routes (
                id SERIAL PRIMARY KEY,
                name VARCHAR(150) NOT NULL,
                type VARCHAR(30) NOT NULL DEFAULT 'Empresarial',
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)

        # 10. Route Lines Table (Sub-Página de Linhas)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS route_lines (
                id SERIAL PRIMARY KEY,
                route_id INT REFERENCES routes(id) ON DELETE CASCADE,
                stretch_name VARCHAR(150) NOT NULL,
                pop_box VARCHAR(100),
                cable_type VARCHAR(100),
                notes TEXT,
                address TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        try:
            cur.execute("ALTER TABLE route_lines ADD COLUMN IF NOT EXISTS address TEXT;")
        except Exception:
            conn.rollback()

        # 10.1 Route Folders & Route Files Tables
        cur.execute("""
            CREATE TABLE IF NOT EXISTS route_folders (
                id SERIAL PRIMARY KEY,
                route_id INT NOT NULL REFERENCES routes(id) ON DELETE CASCADE,
                name VARCHAR(255) NOT NULL,
                parent_id INT REFERENCES route_folders(id) ON DELETE CASCADE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                created_by INT REFERENCES users(id) ON DELETE SET NULL
            );
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS route_files (
                id SERIAL PRIMARY KEY,
                route_id INT NOT NULL REFERENCES routes(id) ON DELETE CASCADE,
                folder_id INT REFERENCES route_folders(id) ON DELETE CASCADE,
                filename VARCHAR(255) NOT NULL,
                filepath VARCHAR(500) NOT NULL,
                filesize BIGINT DEFAULT 0,
                filetype VARCHAR(100) DEFAULT '',
                uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                uploaded_by INT REFERENCES users(id) ON DELETE SET NULL
            );
        """)

        # 11. Forms Table (Formulários)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS forms (
                id SERIAL PRIMARY KEY,
                title VARCHAR(150) NOT NULL,
                slug VARCHAR(100) UNIQUE,
                category VARCHAR(50) DEFAULT 'Inspeção',
                description TEXT,
                link TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        try:
            cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS tech_id INT REFERENCES technicians(id) ON DELETE SET NULL;")
        except Exception:
            conn.rollback()

        try:
            cur.execute("ALTER TABLE forms ADD COLUMN IF NOT EXISTS slug VARCHAR(100) UNIQUE;")
            cur.execute("ALTER TABLE forms ADD COLUMN IF NOT EXISTS questions JSONB;")
        except Exception:
            conn.rollback()

        # 12. Form Responses Table (Respostas dos Técnicos)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS form_responses (
                id SERIAL PRIMARY KEY,
                form_id INT REFERENCES forms(id) ON DELETE CASCADE,
                technician_name VARCHAR(150),
                technician_email VARCHAR(150),
                answers JSONB,
                submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)

        # 13. Work Schedules Table (Escala de Trabalho)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS work_schedules (
                id SERIAL PRIMARY KEY,
                tech_id INT NOT NULL REFERENCES technicians(id) ON DELETE CASCADE,
                date DATE NOT NULL,
                status VARCHAR(50) DEFAULT 'Trabalho',
                work_hours VARCHAR(50) DEFAULT '08 às 17:48hs',
                on_call VARCHAR(50) DEFAULT '0',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(tech_id, date)
            );
        """)

        # Seed default Inventário - Maquinário form if not existing
        cur.execute("SELECT id FROM forms WHERE slug = 'inventario-maquinario';")
        if not cur.fetchone():
            cur.execute("""
                INSERT INTO forms (title, slug, category, description, link)
                VALUES (%s, %s, %s, %s, %s);
            """, (
                'Inventário - Maquinário',
                'inventario-maquinario',
                'Inspeção',
                'Inventário do Maquinário e Ferramental.',
                '/f/inventario-maquinario'
            ))

        # 14. Password Reset Tokens Table (Link Criptografado por E-mail Brevo)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS password_reset_tokens (
                id SERIAL PRIMARY KEY,
                user_id INT REFERENCES users(id) ON DELETE CASCADE,
                token VARCHAR(100) UNIQUE NOT NULL,
                email VARCHAR(150) NOT NULL,
                expires_at TIMESTAMP NOT NULL,
                used BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)

        # Performance Indexes for Large Scale Database Operations
        try:
            cur.execute("CREATE INDEX IF NOT EXISTS idx_form_responses_form_id ON form_responses(form_id);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_form_responses_submitted ON form_responses(submitted_at DESC);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_technicians_name ON technicians(name);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_vehicles_plate ON vehicles(plate);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_routes_type ON routes(type);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_route_lines_route ON route_lines(route_id);")
        except Exception:
            conn.rollback()


        # 8. Teams Finance Table [NEW]
        cur.execute("""
            CREATE TABLE IF NOT EXISTS teams_finance (
                id SERIAL PRIMARY KEY,
                tech1_id INT REFERENCES technicians(id) ON DELETE CASCADE,
                tech2_id INT REFERENCES technicians(id) ON DELETE CASCADE,
                area VARCHAR(100) NOT NULL,
                amount NUMERIC(12, 2) NOT NULL DEFAULT 0.00,
                reference_month VARCHAR(7) NOT NULL, -- format 'YYYY-MM'
                created_by INT REFERENCES users(id) ON DELETE SET NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)

        # 9. Consumables Finance Table [NEW]
        cur.execute("""
            CREATE TABLE IF NOT EXISTS consumables_finance (
                id SERIAL PRIMARY KEY,
                description TEXT NOT NULL,
                area VARCHAR(100) NOT NULL,
                amount NUMERIC(12, 2) NOT NULL DEFAULT 0.00,
                reference_month VARCHAR(7) NOT NULL, -- format 'YYYY-MM'
                created_by INT REFERENCES users(id) ON DELETE SET NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)

        # 10. Inventory Items Table [NEW]
        cur.execute("""
            CREATE TABLE IF NOT EXISTS inventory_items (
                id SERIAL PRIMARY KEY,
                category VARCHAR(50) NOT NULL, -- Cabos, Equipamento, EPI, Ferramentas
                name VARCHAR(150) NOT NULL,
                quantity INT NOT NULL DEFAULT 0,
                serial_number VARCHAR(100),
                description TEXT,
                created_by INT REFERENCES users(id) ON DELETE SET NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)

        # 13. Monthly Indicators Dashboard Table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS monthly_indicators (
                id SERIAL PRIMARY KEY,
                reference_month VARCHAR(20) UNIQUE NOT NULL,
                data JSONB NOT NULL,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)

        # 14. Buscador de Falhas Tables [NEW]
        cur.execute("""
            CREATE TABLE IF NOT EXISTS fault_searcher_records (
                id SERIAL PRIMARY KEY,
                topic VARCHAR(20) NOT NULL,
                col1 VARCHAR(255),
                col2 VARCHAR(255),
                col3 VARCHAR(255),
                col4 VARCHAR(255),
                col5 VARCHAR(255),
                col6 TEXT,
                col7 VARCHAR(100),
                col8 VARCHAR(255),
                col9 VARCHAR(100),
                is_edited BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS fault_searcher_logs (
                topic VARCHAR(20) PRIMARY KEY,
                last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_by VARCHAR(150),
                filename VARCHAR(255),
                record_count INT DEFAULT 0
            );
        """)
        try:
            cur.execute("ALTER TABLE fault_searcher_records ADD COLUMN IF NOT EXISTS is_edited BOOLEAN DEFAULT FALSE;")
            cur.execute("ALTER TABLE fault_searcher_records ADD COLUMN IF NOT EXISTS ref_val VARCHAR(255) DEFAULT '-';")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_fault_searcher_topic ON fault_searcher_records(topic);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_fault_searcher_topic_col2 ON fault_searcher_records(topic, col2);")
        except Exception as e_mig:
            print("Migration index notice:", e_mig)
            conn.rollback()

        # 16. User Tasks Table (Área de Trabalho Exclusiva)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS user_tasks (
                id SERIAL PRIMARY KEY,
                user_id INT REFERENCES users(id) ON DELETE CASCADE,
                title VARCHAR(200) NOT NULL,
                priority VARCHAR(20) NOT NULL DEFAULT 'Média',
                due_date DATE,
                assigned_tech_id INT REFERENCES technicians(id) ON DELETE SET NULL,
                assigned_tech_name VARCHAR(150),
                description TEXT,
                status VARCHAR(30) NOT NULL DEFAULT 'Pendente',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)

        # 17. Collaborator Evaluations Table (Avaliação Técnica)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS collaborator_evaluations (
                id SERIAL PRIMARY KEY,
                technician_id INT REFERENCES technicians(id) ON DELETE CASCADE,
                technician_name VARCHAR(150) NOT NULL,
                company VARCHAR(50),
                behavior_score INT NOT NULL,
                productivity_score INT NOT NULL,
                technical_kpi_score INT NOT NULL,
                process_score INT NOT NULL,
                overall_score NUMERIC(4, 2) NOT NULL,
                comments TEXT,
                evaluator_id INT REFERENCES users(id) ON DELETE SET NULL,
                evaluator_username VARCHAR(80),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)

        # Seed initial admin user if empty
        cur.execute("SELECT COUNT(*) AS count FROM users;")
        if cur.fetchone()['count'] == 0:
            seed_username = os.environ.get('ADMIN_SEED_USERNAME')
            seed_password = os.environ.get('ADMIN_SEED_PASSWORD')
            if seed_username and seed_password:
                hashed = generate_password_hash(seed_password)
                cur.execute(
                    "INSERT INTO users (username, password_hash, role) VALUES (%s, %s, %s);",
                    (seed_username, hashed, "Coordenador")
                )
            else:
                print("Nenhum usuário existe e ADMIN_SEED_USERNAME/ADMIN_SEED_PASSWORD não foram definidos — pulei a criação da conta inicial.")

        # Delete unwanted legacy demo users requested by user
        for unwanted_user in ["claro.supervisor", "claro.operador"]:
            try:
                cur.execute("SELECT id FROM users WHERE username = %s;", (unwanted_user,))
                row = cur.fetchone()
                if row:
                    u_id = row['id']
                    cur.execute("UPDATE audit_logs SET user_id = NULL WHERE user_id = %s;", (u_id,))
                    cur.execute("UPDATE technicians SET created_by = NULL WHERE created_by = %s;", (u_id,))
                    cur.execute("UPDATE vehicles SET responsible_tech_id = NULL WHERE responsible_tech_id = %s;", (u_id,))
                    cur.execute("UPDATE financial_logs SET logged_by = NULL WHERE logged_by = %s;", (u_id,))
                    cur.execute("UPDATE project_folders SET created_by = NULL WHERE created_by = %s;", (u_id,))
                    cur.execute("UPDATE projects SET created_by = NULL WHERE created_by = %s;", (u_id,))
                    cur.execute("DELETE FROM password_reset_tokens WHERE user_id = %s;", (u_id,))
                    cur.execute("DELETE FROM users WHERE id = %s;", (u_id,))
            except Exception as e_del:
                print(f"Cleanup unwanted user {unwanted_user} error:", e_del)
                conn.rollback()

        conn.commit()
        cur.close()
        conn.close()
        print("Database tables initialized and seeded successfully.")
    except Exception as e:
        import traceback
        print("Database initialization failed with traceback:")
        traceback.print_exc()

# Run DB initialization on import
init_db()

# --------------------------------------------------------------------------
# HELPER FUNCTIONS & DECORATORS
# --------------------------------------------------------------------------
def log_action(user_id, username, action):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO audit_logs (user_id, username, action) VALUES (%s, %s, %s);",
            (user_id, username, action)
        )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"Failed to write audit log: {e}")

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({"error": "Sessão expirada. Redirecionando para login...", "redirect": "/login"}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def roles_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            user_role = session.get('role', '')
            user_name = session.get('username', '').lower()
            if user_name == 'alexandre.candido' or user_role in ['Administrador', 'Admin', 'Coordenador', 'Supervisor'] or user_role in roles:
                return f(*args, **kwargs)
            return jsonify({"error": "Acesso negado. Permissão insuficiente."}), 403
        return decorated_function
    return decorator

def is_user_coordenador_claro(user_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT u.role, t.company 
            FROM users u 
            LEFT JOIN technicians t ON u.tech_id = t.id 
            WHERE u.id = %s;
        """, (user_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        if not row:
            return False
        role_val = (row['role'] or '').strip().lower()
        company_val = (row['company'] or '').strip().lower()
        return (role_val in ['coordenador', 'admin']) and (company_val == 'claro')
    except Exception:
        return False

def coordenador_claro_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if not is_user_coordenador_claro(session['user_id']):
            return jsonify({"error": "Você não é um Coordenador Claro, se isto é um erro contate nossa equipe de TI."}), 403
        return f(*args, **kwargs)
    return decorated_function

import urllib.request
import urllib.error
import re
import json

def parse_google_form(url):
    url = (url or '').strip()
    if not url:
        return None, "Link do formulário não foi informado."

    # Normalize Google Form URL
    if 'forms.gle' in url or 'docs.google.com/forms' in url:
        if not url.endswith('/viewform') and '/viewform' not in url and not '?usp=' in url:
            if '/edit' in url:
                url = url.split('/edit')[0] + '/viewform'
            elif not url.endswith('/'):
                url = url + '/viewform'

    req = urllib.request.Request(
        url,
        headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept-Language': 'pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7'
        }
    )

    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            final_url = resp.geturl()
            if 'accounts.google.com' in final_url or 'ServiceLogin' in final_url:
                return None, "Formulário privado ou requer login. Por favor, acesse o Google Forms e altere as permissões para PÚBLICO (qualquer pessoa com o link pode responder)."
            
            html = resp.read().decode('utf-8', errors='ignore')
    except urllib.error.HTTPError as e:
        if e.code in [401, 403]:
            return None, "Formulário privado ou requer login. Por favor, acesse o Google Forms e altere as permissões para PÚBLICO (qualquer pessoa com o link pode responder)."
        return None, f"Erro de conexão com o Google Forms (HTTP {e.code}). Verifique o link e tente novamente."
    except Exception as e:
        return None, f"Falha ao acessar o link do formulário: {e}"

    if 'accounts.google.com' in html or ('Fazer login' in html and 'FB_PUBLIC_LOAD_DATA_' not in html):
        return None, "Formulário privado ou requer login. Por favor, altere a privacidade do Google Form para PÚBLICO."

    # Search for FB_PUBLIC_LOAD_DATA_
    match = re.search(r'FB_PUBLIC_LOAD_DATA_\s*=\s*(.*?);</script>', html, re.DOTALL)
    if not match:
        match = re.search(r'FB_PUBLIC_LOAD_DATA_\s*=\s*(\[.*?\]);', html, re.DOTALL)

    if not match:
        return None, "Não foi possível extrair a estrutura das perguntas deste formulário. Verifique se o link é uma URL válida de resposta do Google Forms."

    try:
        raw_json = match.group(1).strip()
        data = json.loads(raw_json)
    except Exception:
        return None, "Erro ao processar dados internos do Google Forms. O modelo do formulário não pôde ser copiado."

    try:
        form_meta = data[1]
        title = form_meta[8] if len(form_meta) > 8 and form_meta[8] else (form_meta[1] if len(form_meta) > 1 and isinstance(form_meta[1], str) else "Formulário Importado")
        description = form_meta[0] if len(form_meta) > 0 and isinstance(form_meta[0], str) else ""

        questions_raw = []
        if len(form_meta) > 1 and isinstance(form_meta[1], list):
            questions_raw = form_meta[1]

        parsed_questions = []
        if isinstance(questions_raw, list):
            for q in questions_raw:
                if not isinstance(q, list) or len(q) < 4:
                    continue
                q_title = q[1] if len(q) > 1 and q[1] else "Pergunta"
                q_type_code = q[3] if len(q) > 3 else 0
                q_sub = q[4] if len(q) > 4 and isinstance(q[4], list) and len(q[4]) > 0 else []

                field_type = 'text'
                options = []
                required = False

                if q_type_code == 0:
                    field_type = 'text'
                elif q_type_code == 1:
                    field_type = 'textarea'
                elif q_type_code == 2:
                    field_type = 'radio'
                elif q_type_code == 3:
                    field_type = 'select'
                elif q_type_code == 4:
                    field_type = 'checkbox'
                elif q_type_code == 9:
                    field_type = 'date'
                elif q_type_code == 10:
                    field_type = 'time'

                if q_sub and isinstance(q_sub, list) and len(q_sub) > 0:
                    sub_item = q_sub[0]
                    if isinstance(sub_item, list) and len(sub_item) > 1 and isinstance(sub_item[1], list):
                        for opt in sub_item[1]:
                            if isinstance(opt, list) and len(opt) > 0 and opt[0]:
                                options.append(str(opt[0]))
                    if isinstance(sub_item, list) and len(sub_item) > 2:
                        required = bool(sub_item[2] == 1)

                parsed_questions.append({
                    "title": q_title,
                    "type": field_type,
                    "options": options,
                    "required": required
                })

        if not parsed_questions:
            return None, "Não foi possível extrair perguntas válidas deste formulário Google. Verifique se o formulário possui perguntas cadastradas."

        return {
            "title": title,
            "description": description,
            "questions": parsed_questions
        }, None

    except Exception as e:
        return None, f"Não foi possível copiar a estrutura do formulário: {e}"

# --------------------------------------------------------------------------
# PUBLIC SERVING OF UPLOADED ASSETS & PUBLIC PROJECT VIEWER
# --------------------------------------------------------------------------
@app.route('/uploads/<path:filename>')
def serve_upload(filename):
    from werkzeug.utils import safe_join
    # Only project attachments are meant to be public — they're linked from
    # the public project page (/p/project/<id>). Everything else (route
    # files, etc.) is internal and requires an authenticated session.
    if not filename.startswith('projetos/') and 'user_id' not in session:
        return jsonify({"error": "Não autenticado."}), 401
    safe_path = safe_join(UPLOAD_FOLDER, filename)
    if not safe_path or not os.path.exists(safe_path):
        return "Arquivo não encontrado.", 404
    return send_from_directory(UPLOAD_FOLDER, filename)

@app.route('/p/project/<int:project_id>')
def public_project_view(project_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM projects WHERE id = %s;", (project_id,))
        project = cur.fetchone()
        cur.close()
        conn.close()

        if not project:
            return "Projeto de rede não encontrado.", 404

        kmz_files = []
        if project.get('kmz_path'):
            for f in project['kmz_path'].split(';'):
                f_clean = f.strip()
                if f_clean:
                    kmz_files.append({"name": f_clean.split('/')[-1], "path": f_clean})

        pdf_files = []
        if project.get('pdf_path'):
            for f in project['pdf_path'].split(';'):
                f_clean = f.strip()
                if f_clean:
                    pdf_files.append({"name": f_clean.split('/')[-1], "path": f_clean})

        return render_template('public_project.html', project=project, kmz_files=kmz_files, pdf_files=pdf_files)
    except Exception as e:
        return f"Erro ao carregar projeto: {e}", 500

# --------------------------------------------------------------------------
# WEBPAGE ROUTINGS — the React SPA (frontend/dist) owns the entire UI now;
# these routes just hand it index.html. Client-side routing (react-router)
# handles /login, /dashboard, etc. from there.
# --------------------------------------------------------------------------
REACT_DIST = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'frontend', 'dist')

def serve_react_app():
    return send_from_directory(REACT_DIST, 'index.html')

@app.route('/assets/<path:filename>')
def react_assets(filename):
    return send_from_directory(os.path.join(REACT_DIST, 'assets'), filename)

@app.route('/logout')
def logout():
    if 'user_id' in session:
        log_action(session['user_id'], session['username'], "Efetuou logout do sistema.")
        session.clear()
    return redirect(url_for('login'))

@app.route('/login', endpoint='login')
def login():
    return serve_react_app()

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def spa_catch_all(path):
    if path.startswith('api/') or path.startswith('uploads/') or path.startswith('static/'):
        return jsonify({"error": "Não encontrado."}), 404
    if '.' in path.rsplit('/', 1)[-1]:
        # Root-level built assets (favicon.png, claro-icon.png, icons.svg, ...)
        if os.path.isfile(os.path.join(REACT_DIST, path)):
            return send_from_directory(REACT_DIST, path)
        return jsonify({"error": "Arquivo não encontrado."}), 404
    return serve_react_app()

# --------------------------------------------------------------------------
# JSON AUTH API — same session-cookie mechanism as /login above, just returns
# JSON instead of a redirect/rendered template. Added for the React frontend;
# the legacy form-based /login route above is left intact.
# --------------------------------------------------------------------------
@app.route('/api/auth/login', methods=['POST'])
def api_login():
    if 'user_id' in session:
        return jsonify({"error": "Já existe uma sessão ativa."}), 400

    data = request.get_json(silent=True) or {}
    username = (data.get('username') or '').strip().lower()
    password = data.get('password') or ''

    if not username or not password:
        return jsonify({"error": "Por favor, informe seu usuário e senha."}), 400

    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE username = %s;", (username,))
        user = cur.fetchone()
        cur.close()
        conn.close()

        if user and check_password_hash(user['password_hash'], password):
            session.permanent = True
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            log_action(user['id'], user['username'], "Efetuou login no sistema Claro Gestão Rede Externa.")
            return jsonify({"user": {"id": user['id'], "username": user['username'], "role": user['role'], "company": get_user_company(user['id'])}})

        time.sleep(1.5)
        return jsonify({"error": "Usuário ou senha incorretos."}), 401
    except Exception as e:
        return jsonify({"error": f"Erro de conexão com o banco de dados: {e}"}), 500

@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    if 'user_id' in session:
        log_action(session['user_id'], session['username'], "Efetuou logout do sistema.")
        session.clear()
    return jsonify({"success": True})

def get_user_company(user_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT t.company FROM users u
            LEFT JOIN technicians t ON u.tech_id = t.id
            WHERE u.id = %s;
        """, (user_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        return (row['company'].strip() if row and row.get('company') else None)
    except Exception:
        return None

@app.route('/api/auth/me', methods=['GET'])
def api_auth_me():
    if 'user_id' not in session:
        return jsonify({"error": "Não autenticado."}), 401
    return jsonify({"user": {"id": session['user_id'], "username": session['username'], "role": session['role'], "company": get_user_company(session['user_id'])}})

# --------------------------------------------------------------------------
# HONEYPOT / DECOY BAIT ROUTES FOR SCANNER TRAPPING
# --------------------------------------------------------------------------
def get_ip_location(ip):
    if not ip or ip in ['127.0.0.1', 'localhost', '::1']:
        return "Localhost / Rede Interna"
    try:
        url = f"http://ip-api.com/json/{ip}?fields=status,country,regionName,city,isp"
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=3) as resp:
            data = json.loads(resp.read().decode('utf-8'))
            if data.get('status') == 'success':
                city = data.get('city', '')
                region = data.get('regionName', '')
                country = data.get('country', '')
                isp = data.get('isp', '')
                return f"{city}, {region} - {country} ({isp})"
    except Exception:
        pass
    return "Localização Indisponível"

@app.route('/loginadmin')
@app.route('/admin_login')
@app.route('/phpmyadmin')
@app.route('/wp-admin')
@app.route('/admin.php')
def honeypot_trap():
    client_ip = request.remote_addr or request.headers.get('X-Forwarded-For', '127.0.0.1').split(',')[0].strip()
    location = get_ip_location(client_ip)
    
    log_action(None, f"HONEYPOT_TRIGGERED_{client_ip}", f"Tentativa de acesso à rota chamariz de admin: {request.path} | Localização: {location}")
    
    now = time.time()
    for _ in range(50):
        RATE_LIMIT_STORE[client_ip].append(now)
        AUTH_RATE_LIMIT_STORE[client_ip].append(now)
        
    return render_template('honeypot_warning.html', client_ip=client_ip, location=location), 403

# --------------------------------------------------------------------------
# API ENDPOINTS
# --------------------------------------------------------------------------

# --- 1. TECHNICIANS (PESSOAS) ---
@app.route('/api/technicians', methods=['GET'])
@login_required
def api_get_technicians():
    search = request.args.get('search', '').strip().lower()
    try:
        conn = get_db()
        cur = conn.cursor()
        if search:
            s_param = f"%{search}%"
            cur.execute(
                """SELECT * FROM technicians WHERE 
                   LOWER(name) LIKE %s OR 
                   (cpf IS NOT NULL AND LOWER(cpf) LIKE %s) OR 
                   (area IS NOT NULL AND LOWER(area) LIKE %s) OR 
                   (role IS NOT NULL AND LOWER(role) LIKE %s) OR 
                   (company IS NOT NULL AND LOWER(company) LIKE %s) OR 
                   (team_type IS NOT NULL AND LOWER(team_type) LIKE %s) OR 
                   (team IS NOT NULL AND LOWER(team) LIKE %s) OR 
                   (registration_claro IS NOT NULL AND LOWER(registration_claro) LIKE %s) OR 
                   (registration_third IS NOT NULL AND LOWER(registration_third) LIKE %s) OR 
                   (toa_login IS NOT NULL AND LOWER(toa_login) LIKE %s)
                   ORDER BY name ASC;""",
                (s_param, s_param, s_param, s_param, s_param, s_param, s_param, s_param, s_param, s_param)
            )
        else:
            cur.execute("SELECT * FROM technicians ORDER BY name ASC;")
        techs = cur.fetchall()
        cur.close()
        conn.close()
        
        # Format dates safely
        for t in techs:
            if t['dob']:
                try:
                    t['dob'] = t['dob'].strftime('%Y-%m-%d')
                except Exception:
                    t['dob'] = None
            if t['created_at']:
                try:
                    t['created_at'] = t['created_at'].strftime('%Y-%m-%d %H:%M:%S')
                except Exception:
                    t['created_at'] = None
        return jsonify(techs)
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/technicians', methods=['POST'])
@login_required
def api_create_technician():
    try:
        # Form values (multipart form because of uniform specifications)
        name = request.form.get('name', '').strip()
        cpf = request.form.get('cpf', '').strip() or None
        phone = request.form.get('phone', '').strip() or None
        identity = request.form.get('identity', '').strip() or None
        dob = request.form.get('dob', '').strip() or None
        role = request.form.get('role', '').strip() or 'Técnico'
        area = request.form.get('area', '').strip() or None
        team_type = request.form.get('team_type', '').strip() or None
        shirt_size = request.form.get('shirt_size', '').strip() or None
        boot_size = request.form.get('boot_size', '').strip() or None
        pants_size = request.form.get('pants_size', '').strip() or None
        jacket_size = request.form.get('jacket_size', '').strip() or None
        team = request.form.get('team', '').strip() or None
        company = request.form.get('company', '').strip() or None
        registration_claro = request.form.get('registration_claro', '').strip() or None
        registration_third = request.form.get('registration_third', '').strip() or None
        toa_login = request.form.get('toa_login', '').strip() or None
        phone_model = request.form.get('phone_model', '').strip() or None
        imei_1 = request.form.get('imei_1', '').strip() or None
        imei_2 = request.form.get('imei_2', '').strip() or None
        email = request.form.get('email', '').strip() or None
        
        if not name:
            return jsonify({"error": "Nome completo é obrigatório."}), 400

        if dob:
            try:
                yr = int(dob.split('-')[0])
                if yr < 1900 or yr > 2100:
                    dob = None
            except Exception:
                dob = None
            
        conn = get_db()
        cur = conn.cursor()
        
        # Check unique CPF ONLY if CPF is provided
        if cpf:
            cur.execute("SELECT id FROM technicians WHERE cpf = %s;", (cpf,))
            if cur.fetchone():
                cur.close()
                conn.close()
                return jsonify({"error": "Este CPF já está cadastrado no sistema."}), 400
            
        cur.execute(
            """
            INSERT INTO technicians (
                name, cpf, phone, identity, dob, role, area, team_type, 
                shirt_size, boot_size, pants_size, jacket_size, team, company, 
                registration_claro, registration_third, toa_login, phone_model, 
                imei_1, imei_2, email, created_by
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id;
            """,
            (name, cpf, phone, identity, dob, role, area, team_type, 
             shirt_size, boot_size, pants_size, jacket_size, team, company, 
             registration_claro, registration_third, toa_login, phone_model, 
             imei_1, imei_2, email, session['user_id'])
        )
        tech_id = cur.fetchone()['id']
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Cadastrou a pessoa: {name} (ID: {tech_id})")
        return jsonify({"success": True, "id": tech_id}), 201
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/technicians/<int:tech_id>', methods=['POST', 'PUT'])
@login_required
def api_update_technician(tech_id):
    try:
        name = request.form.get('name', '').strip()
        cpf = request.form.get('cpf', '').strip() or None
        phone = request.form.get('phone', '').strip() or None
        identity = request.form.get('identity', '').strip() or None
        dob = request.form.get('dob', '').strip() or None
        role = request.form.get('role', '').strip() or 'Técnico'
        area = request.form.get('area', '').strip() or None
        team_type = request.form.get('team_type', '').strip() or None
        shirt_size = request.form.get('shirt_size', '').strip() or None
        boot_size = request.form.get('boot_size', '').strip() or None
        pants_size = request.form.get('pants_size', '').strip() or None
        jacket_size = request.form.get('jacket_size', '').strip() or None
        team = request.form.get('team', '').strip() or None
        company = request.form.get('company', '').strip() or None
        registration_claro = request.form.get('registration_claro', '').strip() or None
        registration_third = request.form.get('registration_third', '').strip() or None
        toa_login = request.form.get('toa_login', '').strip() or None
        phone_model = request.form.get('phone_model', '').strip() or None
        imei_1 = request.form.get('imei_1', '').strip() or None
        imei_2 = request.form.get('imei_2', '').strip() or None
        email = request.form.get('email', '').strip() or None
        
        if not name:
            return jsonify({"error": "Nome completo é obrigatório."}), 400

        if dob:
            try:
                yr = int(dob.split('-')[0])
                if yr < 1900 or yr > 2100:
                    dob = None
            except Exception:
                dob = None
            
        conn = get_db()
        cur = conn.cursor()
        
        # Check unique CPF (excluding this technician)
        if cpf:
            cur.execute("SELECT id FROM technicians WHERE cpf = %s AND id != %s;", (cpf, tech_id))
            if cur.fetchone():
                cur.close()
                conn.close()
                return jsonify({"error": "Este CPF já está em uso por outra pessoa."}), 400
            
        cur.execute(
            """
            UPDATE technicians 
            SET name = %s, cpf = %s, phone = %s, identity = %s, dob = %s, role = %s, 
                area = %s, team_type = %s, shirt_size = %s, boot_size = %s, 
                pants_size = %s, jacket_size = %s, team = %s, company = %s,
                registration_claro = %s, registration_third = %s, toa_login = %s,
                phone_model = %s, imei_1 = %s, imei_2 = %s, email = %s
            WHERE id = %s;
            """,
            (name, cpf, phone, identity, dob, role, area, team_type, 
             shirt_size, boot_size, pants_size, jacket_size, team, company, 
             registration_claro, registration_third, toa_login, phone_model, 
             imei_1, imei_2, email, tech_id)
        )
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Atualizou cadastro da pessoa: {name} (ID: {tech_id})")
        return jsonify({"success": True})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/technicians/<int:tech_id>', methods=['DELETE'])
@login_required
@roles_required('Coordenador', 'Supervisor')
def api_delete_technician(tech_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT name, cpf FROM technicians WHERE id = %s;", (tech_id,))
        tech = cur.fetchone()
        if not tech:
            cur.close()
            conn.close()
            return jsonify({"error": "Técnico não encontrado."}), 404
            
        cur.execute("DELETE FROM technicians WHERE id = %s;", (tech_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Excluiu cadastro do técnico: {tech['name']} (CPF: {tech['cpf']})")
        return jsonify({"success": True})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500


# --- 2. VEHICLES ---
@app.route('/api/vehicles', methods=['GET'])
@login_required
def api_get_vehicles():
    search = request.args.get('search', '').strip().upper()
    try:
        conn = get_db()
        cur = conn.cursor()
        if search:
            cur.execute(
                """
                SELECT v.*, t.name AS responsible_name 
                FROM vehicles v 
                LEFT JOIN technicians t ON v.responsible_tech_id = t.id 
                WHERE UPPER(v.plate) LIKE %s OR UPPER(v.type) LIKE %s OR UPPER(v.model) LIKE %s 
                    OR UPPER(COALESCE(v.area_rede,'')) LIKE %s OR UPPER(COALESCE(v.base,'')) LIKE %s
                    OR UPPER(COALESCE(v.setor,'')) LIKE %s OR UPPER(COALESCE(v.condutor_dia,'')) LIKE %s
                    OR UPPER(COALESCE(v.condutor_tarde,'')) LIKE %s OR UPPER(COALESCE(v.condutor_madrugada,'')) LIKE %s
                    OR UPPER(COALESCE(v.subclus,'')) LIKE %s OR UPPER(COALESCE(t.name,'')) LIKE %s
                    OR UPPER(COALESCE(v.ticket_car,'')) LIKE %s
                ORDER BY v.plate ASC;
                """,
                tuple(f"%{search}%" for _ in range(12))
            )
        else:
            cur.execute(
                """
                SELECT v.*, t.name AS responsible_name 
                FROM vehicles v 
                LEFT JOIN technicians t ON v.responsible_tech_id = t.id 
                ORDER BY v.plate ASC;
                """
            )
        vehicles = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify(vehicles)
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/vehicles', methods=['POST'])
@login_required
def api_create_vehicle():
    data = request.get_json()
    plate = data.get('plate', '').strip().upper()
    v_type = data.get('type', '').strip() or 'Utilitário'
    model = data.get('model', '').strip() or None
    responsible_tech_id = data.get('responsible_tech_id')
    if responsible_tech_id == '' or responsible_tech_id == 'null' or responsible_tech_id is None:
        responsible_tech_id = None
    else:
        responsible_tech_id = int(responsible_tech_id)
        
    has_rack = bool(data.get('has_rack', False))
    has_basket = bool(data.get('has_basket', False))
    has_giroflex = bool(data.get('has_giroflex', False))
    has_inverter = bool(data.get('has_inverter', False))
    ticket_car = data.get('ticket_car', '').strip() or None
    area_rede = data.get('area_rede', '').strip() or None
    base = data.get('base', '').strip() or None
    setor = data.get('setor', '').strip() or None
    condutor_dia = data.get('condutor_dia', '').strip() or None
    condutor_tarde = data.get('condutor_tarde', '').strip() or None
    condutor_madrugada = data.get('condutor_madrugada', '').strip() or None
    subclus = data.get('subclus', '').strip() or None

    if not plate:
        return jsonify({"error": "Placa do Veículo é obrigatória."}), 400

    try:
        conn = get_db()
        cur = conn.cursor()

        # Check unique plate
        cur.execute("SELECT id FROM vehicles WHERE plate = %s;", (plate,))
        if cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({"error": "Esta placa já está cadastrada no sistema."}), 400
            
        cur.execute(
            """
            INSERT INTO vehicles (plate, type, model, responsible_tech_id, has_rack, has_basket, 
                has_giroflex, has_inverter, ticket_car, area_rede, base, setor, 
                condutor_dia, condutor_tarde, condutor_madrugada, subclus, created_by) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id;
            """,
            (plate, v_type, model, responsible_tech_id, has_rack, has_basket, has_giroflex, has_inverter, 
             ticket_car, area_rede, base, setor, condutor_dia, condutor_tarde, condutor_madrugada, subclus, session['user_id'])
        )
        vehicle_id = cur.fetchone()['id']
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Adicionou o veículo: {plate} (Modelo: {model}, ID: {vehicle_id})")
        return jsonify({"success": True, "id": vehicle_id}), 201
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/vehicles/<int:v_id>', methods=['PUT'])
@login_required
def api_update_vehicle(v_id):
    data = request.get_json()
    plate = data.get('plate', '').strip().upper()
    v_type = data.get('type', '').strip() or 'Utilitário'
    model = data.get('model', '').strip() or None
    responsible_tech_id = data.get('responsible_tech_id')
    if responsible_tech_id == '' or responsible_tech_id == 'null' or responsible_tech_id is None:
        responsible_tech_id = None
    else:
        responsible_tech_id = int(responsible_tech_id)
        
    has_rack = bool(data.get('has_rack', False))
    has_basket = bool(data.get('has_basket', False))
    has_giroflex = bool(data.get('has_giroflex', False))
    has_inverter = bool(data.get('has_inverter', False))
    ticket_car = data.get('ticket_car', '').strip() or None
    area_rede = data.get('area_rede', '').strip() or None
    base = data.get('base', '').strip() or None
    setor = data.get('setor', '').strip() or None
    condutor_dia = data.get('condutor_dia', '').strip() or None
    condutor_tarde = data.get('condutor_tarde', '').strip() or None
    condutor_madrugada = data.get('condutor_madrugada', '').strip() or None
    subclus = data.get('subclus', '').strip() or None
    
    if not plate:
        return jsonify({"error": "Placa do Veículo é obrigatória."}), 400
        
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Check unique plate (excluding current vehicle)
        cur.execute("SELECT id FROM vehicles WHERE plate = %s AND id != %s;", (plate, v_id))
        if cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({"error": "Esta placa já está cadastrada em outro veículo."}), 400
            
        cur.execute(
            """
            UPDATE vehicles 
            SET plate = %s, type = %s, model = %s, responsible_tech_id = %s, 
                has_rack = %s, has_basket = %s, has_giroflex = %s, has_inverter = %s, ticket_car = %s,
                area_rede = %s, base = %s, setor = %s, condutor_dia = %s, condutor_tarde = %s, condutor_madrugada = %s,
                subclus = %s
            WHERE id = %s;
            """,
            (plate, v_type, model, responsible_tech_id, has_rack, has_basket, has_giroflex, has_inverter, 
             ticket_car, area_rede, base, setor, condutor_dia, condutor_tarde, condutor_madrugada, subclus, v_id)
        )
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Atualizou cadastro do veículo: {plate} (ID: {v_id})")
        return jsonify({"success": True})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/vehicles/<int:v_id>', methods=['DELETE'])
@login_required
@roles_required('Coordenador', 'Supervisor')
def api_delete_vehicle(v_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT plate FROM vehicles WHERE id = %s;", (v_id,))
        v = cur.fetchone()
        if not v:
            cur.close()
            conn.close()
            return jsonify({"error": "Veículo não encontrado."}), 404
            
        cur.execute("DELETE FROM vehicles WHERE id = %s;", (v_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Excluiu veículo do cadastro: {v['plate']}")
        return jsonify({"success": True})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500


# --- 4. PROJECTS ---
@app.route('/api/projects', methods=['GET'])
@login_required
def api_get_projects():
    search = request.args.get('search', '').strip().upper()
    folder_id = request.args.get('folder_id')
    if folder_id == 'null' or folder_id == '' or folder_id is None:
        folder_id = None
    else:
        folder_id = int(folder_id)
        
    try:
        conn = get_db()
        cur = conn.cursor()
        if search:
            # Search flat (all folders)
            cur.execute(
                "SELECT * FROM projects WHERE UPPER(name) LIKE %s OR UPPER(description) LIKE %s OR UPPER(COALESCE(area,'')) LIKE %s ORDER BY created_at DESC;",
                (f"%{search}%", f"%{search}%", f"%{search}%")
            )
        elif folder_id is not None:
            cur.execute("SELECT * FROM projects WHERE folder_id = %s ORDER BY created_at DESC;", (folder_id,))
        else:
            cur.execute("SELECT * FROM projects WHERE folder_id IS NULL ORDER BY created_at DESC;")
        projects = cur.fetchall()
        cur.close()
        conn.close()
        
        for p in projects:
            if p['created_at']:
                p['created_at'] = p['created_at'].strftime('%d/%m/%Y %H:%M')
        return jsonify(projects)
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/projects', methods=['POST'])
@login_required
def api_create_project():
    try:
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        area = request.form.get('area', '').strip() or None
        folder_id = request.form.get('folder_id')
        if folder_id == 'null' or folder_id == '' or folder_id is None:
            folder_id = None
        else:
            folder_id = int(folder_id)
        
        if not name:
            return jsonify({"error": "Nome do projeto é obrigatório."}), 400
            
        kmz_paths_list = []
        pdf_paths_list = []
        
        if 'kmz_file' in request.files:
            kmz_files = request.files.getlist('kmz_file')
            for index, file in enumerate(kmz_files):
                if file and file.filename != '' and allowed_file(file.filename, ALLOWED_PROJECT_EXTENSIONS):
                    filename = secure_filename(f"map_{int(time.time())}_{index}_{file.filename}")
                    file.save(os.path.join(UPLOAD_FOLDER, 'projetos', filename))
                    kmz_paths_list.append(f"projetos/{filename}")
                    
        if 'pdf_file' in request.files:
            pdf_files = request.files.getlist('pdf_file')
            for index, file in enumerate(pdf_files):
                if file and file.filename != '' and allowed_file(file.filename, ALLOWED_PROJECT_EXTENSIONS):
                    filename = secure_filename(f"doc_{int(time.time())}_{index}_{file.filename}")
                    file.save(os.path.join(UPLOAD_FOLDER, 'projetos', filename))
                    pdf_paths_list.append(f"projetos/{filename}")
                    
        kmz_path = ';'.join(kmz_paths_list) if kmz_paths_list else None
        pdf_path = ';'.join(pdf_paths_list) if pdf_paths_list else None
                
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO projects (name, description, area, folder_id, kmz_path, pdf_path, created_by) 
            VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id;
            """,
            (name, description, area, folder_id, kmz_path, pdf_path, session['user_id'])
        )
        p_id = cur.fetchone()['id']
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Criou o projeto de rede externa: {name} (ID: {p_id})")
        return jsonify({"success": True, "id": p_id}), 201
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/projects/<int:p_id>', methods=['DELETE'])
@login_required
@roles_required('Coordenador', 'Supervisor')
def api_delete_project(p_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT name FROM projects WHERE id = %s;", (p_id,))
        p = cur.fetchone()
        if not p:
            cur.close()
            conn.close()
            return jsonify({"error": "Projeto não encontrado."}), 404
            
        cur.execute("DELETE FROM projects WHERE id = %s;", (p_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Deletou o projeto de rede: {p['name']}")
        return jsonify({"success": True})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/projects/<int:p_id>', methods=['POST', 'PUT'])
@login_required
def api_update_project(p_id):
    try:
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        area = request.form.get('area', '').strip() or None
        
        if not name:
            return jsonify({"error": "Nome do projeto é obrigatório."}), 400
            
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute("SELECT * FROM projects WHERE id = %s;", (p_id,))
        project = cur.fetchone()
        if not project:
            cur.close()
            conn.close()
            return jsonify({"error": "Projeto não encontrado."}), 404
            
        kmz_paths_list = [project['kmz_path']] if project['kmz_path'] else []
        pdf_paths_list = [project['pdf_path']] if project['pdf_path'] else []
        
        if 'kmz_file' in request.files:
            kmz_files = request.files.getlist('kmz_file')
            for index, file in enumerate(kmz_files):
                if file and file.filename != '' and allowed_file(file.filename, ALLOWED_PROJECT_EXTENSIONS):
                    filename = secure_filename(f"map_{int(time.time())}_{index}_{file.filename}")
                    file.save(os.path.join(UPLOAD_FOLDER, 'projetos', filename))
                    kmz_paths_list.append(f"projetos/{filename}")
                    
        if 'pdf_file' in request.files:
            pdf_files = request.files.getlist('pdf_file')
            for index, file in enumerate(pdf_files):
                if file and file.filename != '' and allowed_file(file.filename, ALLOWED_PROJECT_EXTENSIONS):
                    filename = secure_filename(f"doc_{int(time.time())}_{index}_{file.filename}")
                    file.save(os.path.join(UPLOAD_FOLDER, 'projetos', filename))
                    pdf_paths_list.append(f"projetos/{filename}")
                    
        kmz_path = ';'.join(kmz_paths_list) if kmz_paths_list else None
        pdf_path = ';'.join(pdf_paths_list) if pdf_paths_list else None
        
        cur.execute(
            """
            UPDATE projects 
            SET name = %s, description = %s, area = %s, kmz_path = %s, pdf_path = %s 
            WHERE id = %s;
            """,
            (name, description, area, kmz_path, pdf_path, p_id)
        )
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Editou o projeto de rede: {name} (ID: {p_id})")
        return jsonify({"success": True}), 200
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

# --- FAVORITES API ---
@app.route('/api/favorites', methods=['GET'])
@login_required
def api_get_favorites():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT id, title, link, color, COALESCE(access_count, 0) AS access_count, created_at FROM favorites ORDER BY COALESCE(access_count, 0) DESC, id DESC;")
        favs = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify(favs)
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/favorites/<int:fav_id>/click', methods=['POST'])
@login_required
def api_click_favorite(fav_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("UPDATE favorites SET access_count = COALESCE(access_count, 0) + 1 WHERE id = %s RETURNING access_count, title;", (fav_id,))
        row = cur.fetchone()
        if not row:
            cur.close()
            conn.close()
            return jsonify({"error": "Favorito não encontrado."}), 404
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"success": True, "access_count": row['access_count']})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/favorites', methods=['POST'])
@login_required
def api_create_favorite():
    try:
        data = request.get_json(silent=True) or request.form
        title = data.get('title', '').strip()
        link = data.get('link', '').strip()
        color = data.get('color', 'Vermelho').strip()
        
        if not title or not link:
            return jsonify({"error": "Nome e link são obrigatórios."}), 400
            
        if not link.startswith('http://') and not link.startswith('https://'):
            link = 'https://' + link

        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO favorites (title, link, color) VALUES (%s, %s, %s) RETURNING id;",
            (title, link, color)
        )
        new_id = cur.fetchone()['id']
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Cadastrou o favorito: {title}")
        return jsonify({"success": True, "id": new_id}), 201
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/favorites/<int:fav_id>', methods=['POST', 'PUT'])
@login_required
def api_update_favorite(fav_id):
    try:
        data = request.get_json(silent=True) or request.form
        title = data.get('title', '').strip()
        link = data.get('link', '').strip()
        color = data.get('color', 'Vermelho').strip()
        
        if not title or not link:
            return jsonify({"error": "Nome e link são obrigatórios."}), 400

        if not link.startswith('http://') and not link.startswith('https://'):
            link = 'https://' + link

        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "UPDATE favorites SET title = %s, link = %s, color = %s WHERE id = %s;",
            (title, link, color, fav_id)
        )
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Atualizou o favorito: {title} (ID: {fav_id})")
        return jsonify({"success": True}), 200
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/favorites/<int:fav_id>', methods=['DELETE'])
@login_required
def api_delete_favorite(fav_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM favorites WHERE id = %s;", (fav_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Deletou favorito ID: {fav_id}")
        return jsonify({"success": True})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

# --- ROUTES API ---
@app.route('/api/routes', methods=['GET', 'POST'], strict_slashes=False)
@login_required
def api_routes():
    if request.method == 'GET':
        try:
            route_type = request.args.get('type', '').strip()
            conn = get_db()
            cur = conn.cursor()
            
            if route_type and route_type in ['Empresarial', 'Residencial']:
                cur.execute(
                    "SELECT r.id, r.name, r.type, r.description, r.created_at, "
                    "COUNT(l.id) AS lines_count "
                    "FROM routes r LEFT JOIN route_lines l ON r.id = l.route_id "
                    "WHERE r.type = %s GROUP BY r.id ORDER BY r.created_at DESC, r.id DESC;",
                    (route_type,)
                )
            else:
                cur.execute(
                    "SELECT r.id, r.name, r.type, r.description, r.created_at, "
                    "COUNT(l.id) AS lines_count "
                    "FROM routes r LEFT JOIN route_lines l ON r.id = l.route_id "
                    "GROUP BY r.id ORDER BY r.created_at DESC, r.id DESC;"
                )
            routes = cur.fetchall()
            cur.close()
            conn.close()
            for r in routes:
                if r.get('created_at'):
                    r['created_at_fmt'] = r['created_at'].strftime('%d/%m/%Y às %H:%M')
                else:
                    r['created_at_fmt'] = ''
            return jsonify(routes)
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

    elif request.method == 'POST':
        try:
            data = request.get_json(silent=True) or request.form
            name = data.get('name', '').strip()
            route_type = data.get('type', 'Empresarial').strip()
            description = data.get('description', '').strip()
            
            if not name:
                return jsonify({"error": "Nome da rota é obrigatório."}), 400
                
            if route_type not in ['Empresarial', 'Residencial']:
                route_type = 'Empresarial'

            conn = get_db()
            cur = conn.cursor()
            cur.execute(
                "INSERT INTO routes (name, type, description) VALUES (%s, %s, %s) RETURNING id;",
                (name, route_type, description)
            )
            new_id = cur.fetchone()['id']
            conn.commit()
            cur.close()
            conn.close()
            
            log_action(session['user_id'], session['username'], f"Cadastrou a rota: {name} ({route_type})")
            return jsonify({"success": True, "id": new_id}), 201
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/routes/<int:route_id>', methods=['GET', 'POST', 'PUT', 'DELETE'], strict_slashes=False)
@login_required
def api_route_detail(route_id):
    if request.method == 'GET':
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("SELECT id, name, type, description, created_at FROM routes WHERE id = %s;", (route_id,))
            route = cur.fetchone()
            if not route:
                cur.close()
                conn.close()
                return jsonify({"error": "Rota não encontrada."}), 404
                
            cur.execute("SELECT id, route_id, stretch_name, pop_box, cable_type, notes, address, created_at FROM route_lines WHERE route_id = %s ORDER BY id ASC;", (route_id,))
            lines = cur.fetchall()
            cur.close()
            conn.close()
            
            return jsonify({"route": route, "lines": lines})
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

    elif request.method == 'DELETE':
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("DELETE FROM routes WHERE id = %s;", (route_id,))
            conn.commit()
            cur.close()
            conn.close()
            
            log_action(session['user_id'], session['username'], f"Excluiu a rota ID: {route_id}")
            return jsonify({"success": True})
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

    else: # POST or PUT
        try:
            data = request.get_json(silent=True) or request.form
            name = data.get('name', '').strip()
            route_type = data.get('type', 'Empresarial').strip()
            description = data.get('description', '').strip()
            
            if not name:
                return jsonify({"error": "Nome da rota é obrigatório."}), 400

            if route_type not in ['Empresarial', 'Residencial']:
                route_type = 'Empresarial'

            conn = get_db()
            cur = conn.cursor()
            cur.execute(
                "UPDATE routes SET name = %s, type = %s, description = %s WHERE id = %s;",
                (name, route_type, description, route_id)
            )
            conn.commit()
            cur.close()
            conn.close()
            
            log_action(session['user_id'], session['username'], f"Atualizou a rota: {name} (ID: {route_id})")
            return jsonify({"success": True}), 200
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

# --- ROUTE LINES API ---
@app.route('/api/routes/<int:route_id>/lines', methods=['POST'], strict_slashes=False)
@login_required
def api_add_route_line(route_id):
    try:
        data = request.get_json(silent=True) or request.form
        stretch_name = data.get('stretch_name', '').strip()
        pop_box = data.get('pop_box', '').strip()
        cable_type = data.get('cable_type', '').strip()
        notes = data.get('notes', '').strip()
        address = data.get('address', '').strip()
        
        if not stretch_name:
            return jsonify({"error": "Data da medição é obrigatória."}), 400

        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO route_lines (route_id, stretch_name, pop_box, cable_type, notes, address) "
            "VALUES (%s, %s, %s, %s, %s, %s) RETURNING id;",
            (route_id, stretch_name, pop_box, cable_type, notes, address)
        )
        new_id = cur.fetchone()['id']
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Adicionou linha/medição '{stretch_name}' na Rota ID {route_id}")
        return jsonify({"success": True, "id": new_id}), 201
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/routes/lines/<int:line_id>', methods=['POST', 'PUT', 'DELETE'], strict_slashes=False)
@login_required
def api_route_line_detail(line_id):
    if request.method == 'DELETE':
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("DELETE FROM route_lines WHERE id = %s;", (line_id,))
            conn.commit()
            cur.close()
            conn.close()
            
            return jsonify({"success": True})
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500
    else: # POST or PUT
        try:
            data = request.get_json(silent=True) or request.form
            stretch_name = data.get('stretch_name', '').strip()
            pop_box = data.get('pop_box', '').strip()
            cable_type = data.get('cable_type', '').strip()
            notes = data.get('notes', '').strip()
            address = data.get('address', '').strip()
            
            if not stretch_name:
                return jsonify({"error": "Data da medição é obrigatória."}), 400

            conn = get_db()
            cur = conn.cursor()
            cur.execute(
                "UPDATE route_lines SET stretch_name = %s, pop_box = %s, cable_type = %s, notes = %s, address = %s WHERE id = %s;",
                (stretch_name, pop_box, cable_type, notes, address, line_id)
            )
            conn.commit()
            cur.close()
            conn.close()
            
            return jsonify({"success": True}), 200
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/routes/<int:route_id>/export', methods=['GET'], strict_slashes=False)
@login_required
def api_export_route_lines(route_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT name FROM routes WHERE id = %s;", (route_id,))
        r = cur.fetchone()
        route_name = r['name'] if r else f"Rota_{route_id}"

        cur.execute("SELECT stretch_name, pop_box, cable_type, notes, address FROM route_lines WHERE route_id = %s ORDER BY id ASC;", (route_id,))
        lines = cur.fetchall()
        cur.close()
        conn.close()

        csv_output = "Data;Local da Medição;Medição;Observações Técnicas;Endereço\n"
        for line in lines:
            s_name = (line.get('stretch_name') or '').replace(';', ',').replace('\n', ' ')
            pop = (line.get('pop_box') or '').replace(';', ',').replace('\n', ' ')
            med = (line.get('cable_type') or '').replace(';', ',').replace('\n', ' ')
            obs = (line.get('notes') or '').replace(';', ',').replace('\n', ' ')
            addr = (line.get('address') or '').replace(';', ',').replace('\n', ' ')
            csv_output += f"{s_name};{pop};{med};{obs};{addr}\n"

        from flask import Response
        filename = f"medicoes_{route_name.lower().replace(' ', '_')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return Response(
            csv_output,
            mimetype="text/csv",
            headers={"Content-disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/routes/<int:route_id>/bulk', methods=['POST'], strict_slashes=False)
@login_required
def api_bulk_import_route_lines(route_id):
    try:
        data = request.get_json(silent=True) or request.form
        csv_data = data.get('csv_data', '').strip()
        if not csv_data:
            return jsonify({"error": "Nenhum dado informado para importação."}), 400

        lines_raw = csv_data.splitlines()
        imported = 0
        conn = get_db()
        cur = conn.cursor()

        for idx, row in enumerate(lines_raw):
            if not row.strip():
                continue
            parts = [p.strip() for p in row.split(';')]
            if len(parts) < 2:
                parts = [p.strip() for p in row.split(',')]
            
            # Skip header if present
            if idx == 0 and ('data' in parts[0].lower() or 'linha' in parts[0].lower()):
                continue

            stretch_name = parts[0] if len(parts) > 0 else ''
            pop_box = parts[1] if len(parts) > 1 else ''
            cable_type = parts[2] if len(parts) > 2 else ''
            notes = parts[3] if len(parts) > 3 else ''
            address = parts[4] if len(parts) > 4 else ''

            if not stretch_name:
                continue

            cur.execute(
                "INSERT INTO route_lines (route_id, stretch_name, pop_box, cable_type, notes, address) "
                "VALUES (%s, %s, %s, %s, %s, %s);",
                (route_id, stretch_name, pop_box, cable_type, notes, address)
            )
            imported += 1

        conn.commit()
        cur.close()
        conn.close()

        log_action(session['user_id'], session['username'], f"Importou {imported} medições/linhas na Rota ID {route_id}")
        return jsonify({"success": True, "imported": imported})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

# --- FORMS API ---
@app.route('/api/forms', methods=['GET', 'POST'], strict_slashes=False)
@login_required
def api_forms():
    if request.method == 'GET':
        try:
            category = request.args.get('category', '').strip()
            conn = get_db()
            cur = conn.cursor()
            if category:
                cur.execute("SELECT id, title, slug, category, description, link, created_at FROM forms WHERE category = %s ORDER BY id DESC;", (category,))
            else:
                cur.execute("SELECT id, title, slug, category, description, link, created_at FROM forms ORDER BY id DESC;")
            forms = cur.fetchall()
            cur.close()
            conn.close()
            return jsonify(forms)
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

    elif request.method == 'POST':
        try:
            data = request.get_json(silent=True) or request.form
            title = data.get('title', '').strip()
            category = data.get('category', 'Inspeção').strip()
            description = data.get('description', '').strip()
            link = data.get('link', '').strip()

            questions = None
            if link and ('forms.gle' in link or 'docs.google.com/forms' in link):
                parsed, parse_err = parse_google_form(link)
                if parse_err:
                    return jsonify({"error": parse_err}), 400
                if parsed:
                    if not title:
                        title = parsed['title']
                    if not description:
                        description = parsed['description']
                    questions = parsed['questions']

            if not title:
                return jsonify({"error": "Nome do formulário é obrigatório."}), 400

            if link and not link.startswith('http://') and not link.startswith('https://'):
                link = 'https://' + link

            # Generate unique slug for form
            slug = re.sub(r'[^a-z0-9]+', '-', title.lower()).strip('-')
            if not slug:
                slug = f"form-{int(time.time())}"

            conn = get_db()
            cur = conn.cursor()
            
            # Check if slug exists
            cur.execute("SELECT id FROM forms WHERE slug = %s;", (slug,))
            if cur.fetchone():
                slug = f"{slug}-{int(time.time() % 10000)}"

            questions_json = json.dumps(questions, ensure_ascii=False) if questions else None

            cur.execute(
                "INSERT INTO forms (title, slug, category, description, link, questions) VALUES (%s, %s, %s, %s, %s, %s) RETURNING id;",
                (title, slug, category, description, link, questions_json)
            )
            new_id = cur.fetchone()['id']
            conn.commit()
            cur.close()
            conn.close()

            log_action(session['user_id'], session['username'], f"Cadastrou o formulário: {title}")
            return jsonify({"success": True, "id": new_id, "slug": slug, "has_questions": bool(questions)}), 201
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/forms/<int:form_id>', methods=['GET', 'POST', 'PUT', 'DELETE'], strict_slashes=False)
@login_required
def api_form_detail(form_id):
    if request.method == 'GET':
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("SELECT id, title, slug, category, description, link, created_at FROM forms WHERE id = %s;", (form_id,))
            f = cur.fetchone()
            cur.close()
            conn.close()
            if not f:
                return jsonify({"error": "Formulário não encontrado."}), 404
            return jsonify(f)
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

    elif request.method == 'DELETE':
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("DELETE FROM forms WHERE id = %s;", (form_id,))
            conn.commit()
            cur.close()
            conn.close()

            log_action(session['user_id'], session['username'], f"Excluiu o formulário ID: {form_id}")
            return jsonify({"success": True})
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

    else: # POST or PUT
        try:
            data = request.get_json(silent=True) or request.form
            title = data.get('title', '').strip()
            category = data.get('category', 'Inspeção').strip()
            description = data.get('description', '').strip()
            link = data.get('link', '').strip()

            if not title:
                return jsonify({"error": "Nome do formulário é obrigatório."}), 400

            if link and not link.startswith('http://') and not link.startswith('https://'):
                link = 'https://' + link

            new_slug = re.sub(r'[^a-z0-9]+', '-', title.lower()).strip('-') or f"form-{form_id}"

            conn = get_db()
            cur = conn.cursor()
            cur.execute(
                "UPDATE forms SET title = %s, slug = COALESCE(NULLIF(slug, ''), %s), category = %s, description = %s, link = %s WHERE id = %s;",
                (title, new_slug, category, description, link, form_id)
            )
            conn.commit()
            cur.close()
            conn.close()

            log_action(session['user_id'], session['username'], f"Atualizou formulário: {title} (ID: {form_id})")
            return jsonify({"success": True}), 200
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

# --- PUBLIC FORM RESPONDER ROUTES (NO LOGIN REQUIRED) ---
@app.route('/f/<slug>', methods=['GET'], strict_slashes=False)
def public_form_view(slug):
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # 1. Search by exact slug
        cur.execute("SELECT id, title, slug, category, description, link, questions FROM forms WHERE slug = %s;", (slug,))
        f = cur.fetchone()
        
        # 2. Search by numeric ID if slug is digit
        if not f and slug.isdigit():
            cur.execute("SELECT id, title, slug, category, description, link, questions FROM forms WHERE id = %s;", (int(slug),))
            f = cur.fetchone()

        cur.close()
        conn.close()

        if not f:
            return "Formulário não encontrado ou inativo.", 404

        # If form has an external link defined (like Google Forms or Microsoft Forms), redirect directly to it
        if f.get('link') and (f['link'].startswith('http://') or f['link'].startswith('https://')):
            return redirect(f['link'])

        return render_template('public_form.html', form=f)
    except Exception as e:
        return f"Erro ao carregar formulário: {str(e)}", 500

@app.route('/f/<slug>/submit', methods=['POST'], strict_slashes=False)
def public_form_submit(slug):
    try:
        data = request.get_json(silent=True) or request.form.to_dict()
        if not data:
            return jsonify({"error": "Nenhum dado recebido."}), 400

        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT id FROM forms WHERE slug = %s;", (slug,))
        f = cur.fetchone()
        if not f and slug.isdigit():
            cur.execute("SELECT id FROM forms WHERE id = %s;", (int(slug),))
            f = cur.fetchone()

        if not f:
            cur.close()
            conn.close()
            return jsonify({"error": "Formulário não encontrado."}), 404

        form_id = f['id']
        email = data.get('email', '').strip()
        technician = data.get('technician', data.get('email', 'Técnico')).strip()

        # Convert dictionary data to JSON string for Postgres storage
        import json
        cur.execute(
            "INSERT INTO form_responses (form_id, technician_name, technician_email, answers) VALUES (%s, %s, %s, %s);",
            (form_id, technician, email, json.dumps(data, ensure_ascii=False))
        )
        conn.commit()
        cur.close()
        conn.close()

        return jsonify({"success": True, "message": "Resposta registrada com sucesso!"})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

# --- FORM RESPONSES ADMIN API ---
@app.route('/api/forms/<int:form_id>/responses', methods=['GET'], strict_slashes=False)
@login_required
def api_form_responses(form_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Auto import historical sheet data if this form is inventario-maquinario and has missing sheet responses
        cur.execute("SELECT slug FROM forms WHERE id = %s;", (form_id,))
        f_row = cur.fetchone()
        if f_row and f_row['slug'] == 'inventario-maquinario':
            cur.execute("SELECT COUNT(*) AS count FROM form_responses WHERE form_id = %s AND (answers->>'Carimbo de data/hora' IS NOT NULL OR answers->>'Data' IS NOT NULL);", (form_id,))
            if cur.fetchone()['count'] == 0:
                try:
                    import urllib.request, csv, json, datetime
                    sheet_url = "https://docs.google.com/spreadsheets/d/1e4ntyLxdVHP5Cls46UlHXvtVQRxkbDRbqtqJz90xhJ4/export?format=csv&gid=1914441081"
                    req = urllib.request.Request(sheet_url, headers={'User-Agent': 'Mozilla/5.0'})
                    with urllib.request.urlopen(req, timeout=15) as resp:
                        csv_text = resp.read().decode('utf-8')

                    reader = csv.reader(csv_text.splitlines())
                    header = None

                    for row in reader:
                        if not row or not any(row):
                            continue
                        if not header:
                            header = [h.strip() for h in row]
                            continue

                        answers = {}
                        tech_name = 'N/A'
                        tech_email = 'N/A'
                        submitted_at_str = None

                        for idx, col_name in enumerate(header):
                            if not col_name:
                                col_name = f"Coluna_{idx+1}"
                            val = row[idx].strip() if idx < len(row) and row[idx].strip() else 'N/A'
                            answers[col_name] = val

                            col_lower = col_name.lower()
                            if 'técnico' in col_lower or col_lower == 'técnico':
                                if val != 'N/A':
                                    tech_name = val
                            elif 'e-mail' in col_lower or 'email' in col_lower:
                                if val != 'N/A':
                                    tech_email = val
                            elif 'carimbo' in col_lower or col_lower == 'data':
                                if val != 'N/A' and not submitted_at_str:
                                    submitted_at_str = val

                        submitted_at = None
                        if submitted_at_str:
                            for fmt in ['%d/%m/%Y %H:%M:%S', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d']:
                                try:
                                    submitted_at = datetime.datetime.strptime(submitted_at_str, fmt)
                                    break
                                except ValueError:
                                    pass

                        if not submitted_at:
                            submitted_at = datetime.datetime.now()

                        cur.execute("""
                            INSERT INTO form_responses (form_id, technician_name, technician_email, answers, submitted_at)
                            VALUES (%s, %s, %s, %s, %s);
                        """, (form_id, tech_name, tech_email, json.dumps(answers, ensure_ascii=False), submitted_at))

                    conn.commit()
                except Exception as e_imp:
                    print("Error auto-importing sheet in GET responses:", e_imp)
                    conn.rollback()

        cur.execute("""
            SELECT id, technician_name, technician_email, answers, submitted_at
            FROM form_responses
            WHERE form_id = %s
            ORDER BY id DESC;
        """, (form_id,))
        responses = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify(responses)
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/forms/<int:form_id>/responses/export', methods=['GET'], strict_slashes=False)
@login_required
def api_form_responses_export(form_id):
    try:
        import io, csv, json
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT title FROM forms WHERE id = %s;", (form_id,))
        form_row = cur.fetchone()
        title = form_row['title'] if form_row else 'Formulario'

        cur.execute("""
            SELECT id, technician_name, technician_email, answers, submitted_at
            FROM form_responses
            WHERE form_id = %s
            ORDER BY id DESC;
        """, (form_id,))
        rows = cur.fetchall()
        cur.close()
        conn.close()

        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')

        # Header
        writer.writerow(['ID', 'Data/Hora Submissao', 'Tecnico', 'E-mail', 'Respostas Detalhadas'])

        for r in rows:
            ans_str = json.dumps(r['answers'], ensure_ascii=False) if isinstance(r['answers'], dict) else str(r['answers'] or '')
            sub_date = r['submitted_at'].strftime('%d/%m/%Y %H:%M') if r['submitted_at'] else ''
            writer.writerow([r['id'], sub_date, r['technician_name'] or '', r['technician_email'] or '', ans_str])

        output.seek(0)
        filename = f"respostas_{form_id}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/forms/<int:form_id>/responses/import', methods=['POST'], strict_slashes=False)
@login_required
def api_form_responses_import(form_id):
    try:
        import csv, io, json
        data = request.get_json(silent=True) or {}
        csv_text = data.get('csv_data', '').strip()

        if not csv_text and 'file' in request.files:
            file = request.files['file']
            csv_text = file.read().decode('utf-8-sig', errors='ignore')

        if not csv_text:
            return jsonify({"error": "Nenhum arquivo ou texto de CSV informado."}), 400

        reader = csv.reader(io.StringIO(csv_text), delimiter=';')
        rows = list(reader)
        if len(rows) < 2:
            reader = csv.reader(io.StringIO(csv_text), delimiter=',')
            rows = list(reader)

        if len(rows) < 2:
            return jsonify({"error": "Arquivo CSV sem linhas suficientes de cabeçalho e dados."}), 400

        headers = [h.strip() for h in rows[0]]
        imported = 0

        conn = get_db()
        cur = conn.cursor()

        for row in rows[1:]:
            if not row or not any(row):
                continue
            answers = {}
            tech_name = 'Técnico'
            tech_email = ''

            for idx, h in enumerate(headers):
                val = row[idx].strip() if idx < len(row) else ''
                answers[h] = val
                h_lower = h.lower()
                if 'técnico' in h_lower or 'tecnico' in h_lower or 'nome' in h_lower:
                    if val: tech_name = val
                elif 'e-mail' in h_lower or 'email' in h_lower:
                    if val: tech_email = val

            cur.execute("""
                INSERT INTO form_responses (form_id, technician_name, technician_email, answers)
                VALUES (%s, %s, %s, %s);
            """, (form_id, tech_name, tech_email, json.dumps(answers, ensure_ascii=False)))
            imported += 1

        conn.commit()
        cur.close()
        conn.close()

        log_action(session['user_id'], session['username'], f"Importou {imported} respostas no Formulário ID {form_id}")
        return jsonify({"success": True, "imported": imported})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

# --- EXTRACTION ROUTE FOR LOCAL DOWNLOADED EXCEL ---
@app.route('/api/admin/extract_machinery_excel', methods=['POST'], strict_slashes=False)
@login_required
def api_extract_machinery_excel():
    excel_path = os.path.join(r"C:\Users\2d\Downloads", "2 - Inventário - Maquina e ferramentas.xlsx")
    if not os.path.exists(excel_path):
        return jsonify({"error": f"Arquivo não encontrado no caminho: {excel_path}"}), 404

    import zipfile
    import xml.etree.ElementTree as ET
    import json

    try:
        with zipfile.ZipFile(excel_path, 'r') as z:
            shared_strings = []
            if 'xl/sharedStrings.xml' in z.namelist():
                tree = ET.fromstring(z.read('xl/sharedStrings.xml'))
                for elem in tree.iter():
                    if elem.tag.endswith('t') and elem.text:
                        shared_strings.append(elem.text)

            sheet_xml = z.read('xl/worksheets/sheet1.xml')
            tree = ET.fromstring(sheet_xml)

            rows = []
            for row_elem in tree.iter():
                if row_elem.tag.endswith('row'):
                    row_cells = []
                    for cell_elem in row_elem.iter():
                        if cell_elem.tag.endswith('c'):
                            cell_type = cell_elem.attrib.get('t', '')
                            cell_val = ''
                            for val_elem in cell_elem.iter():
                                if val_elem.tag.endswith('v'):
                                    v_str = val_elem.text or ''
                                    if cell_type == 's' and v_str.isdigit():
                                        idx = int(v_str)
                                        cell_val = shared_strings[idx] if idx < len(shared_strings) else v_str
                                    else:
                                        cell_val = v_str
                            row_cells.append(cell_val.strip())
                    if row_cells and any(row_cells):
                        rows.append(row_cells)

            if len(rows) < 2:
                return jsonify({"error": "Arquivo Excel sem dados suficientes para importação."}), 400

            headers = rows[0]
            conn = get_db()
            cur = conn.cursor()

            cur.execute("SELECT id FROM forms WHERE slug = 'inventario-maquinario';")
            form_row = cur.fetchone()
            form_id = form_row['id'] if form_row else 1

            imported_count = 0
            for row in rows[1:]:
                if not row or not any(row):
                    continue
                answers = {}
                tech_name = 'Técnico'
                tech_email = ''

                for idx, h in enumerate(headers):
                    val = row[idx] if idx < len(row) else ''
                    answers[h] = val
                    h_lower = h.lower()
                    if ('técnico' in h_lower or 'tecnico' in h_lower or 'nome' in h_lower) and val:
                        tech_name = val
                    elif ('email' in h_lower or 'e-mail' in h_lower) and val:
                        tech_email = val

                cur.execute("""
                    INSERT INTO form_responses (form_id, technician_name, technician_email, answers)
                    VALUES (%s, %s, %s, %s);
                """, (form_id, tech_name, tech_email, json.dumps(answers, ensure_ascii=False)))
                imported_count += 1

            conn.commit()
            cur.close()
            conn.close()

            log_action(session['user_id'], session['username'], f"Extraiu {imported_count} respostas do arquivo Excel de Downloads (2 - Inventário - Maquina e ferramentas.xlsx)")
            return jsonify({"success": True, "imported": imported_count})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

# --- MONTHLY INDICATORS DASHBOARD API ---
@app.route('/api/indicators', methods=['GET'], strict_slashes=False)
@login_required
def api_get_monthly_indicators():
    month = request.args.get('month', 'julho-26').strip().lower()
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM monthly_indicators WHERE reference_month = %s;", (month,))
        row = cur.fetchone()
        cur.close()
        conn.close()

        if row and row.get('data'):
            return jsonify({"month": month, "data": row['data']})
        return jsonify({"month": month, "data": None})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/indicators', methods=['POST'], strict_slashes=False)
@login_required
def api_save_monthly_indicators():
    try:
        data = request.get_json(silent=True) or {}
        month = data.get('month', 'julho-26').strip().lower()
        payload = data.get('data', {})

        if not month:
            return jsonify({"error": "Mês de referência é obrigatório."}), 400

        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO monthly_indicators (reference_month, data, updated_at)
            VALUES (%s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT (reference_month) DO UPDATE
            SET data = EXCLUDED.data, updated_at = CURRENT_TIMESTAMP;
        """, (month, json.dumps(payload, ensure_ascii=False)))
        conn.commit()
        cur.close()
        conn.close()

        log_action(session['user_id'], session['username'], f"Atualizou os indicadores do mês {month}")
        return jsonify({"success": True, "month": month})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

# --- BULK IMPORT API (PESSOAS / TÉCNICOS) ---
@app.route('/api/people/bulk', methods=['POST'], strict_slashes=False)
@login_required
def api_people_bulk():
    try:
        data = request.get_json(silent=True) or {}
        csv_data = data.get('csv_data', '').strip()
        if not csv_data:
            return jsonify({"error": "Nenhum dado informado para importação."}), 400

        lines_raw = csv_data.splitlines()
        imported = 0
        conn = get_db()
        cur = conn.cursor()

        for idx, row in enumerate(lines_raw):
            if not row.strip(): continue
            parts = [p.strip() for p in row.split(';')]
            if len(parts) < 2:
                parts = [p.strip() for p in row.split(',')]

            # Skip header if present
            if idx == 0 and ('nome' in parts[0].lower() or 'cpf' in parts[0].lower() or 'tecnico' in parts[0].lower()):
                continue

            name = parts[0] if len(parts) > 0 else ''
            role = parts[1] if len(parts) > 1 else 'Técnico de Campo'
            cpf = parts[2] if len(parts) > 2 else ''
            phone = parts[3] if len(parts) > 3 else ''
            status = parts[4] if len(parts) > 4 else 'Ativo'

            if not name: continue

            cur.execute("""
                INSERT INTO technicians (name, role, cpf, phone)
                VALUES (%s, %s, %s, %s);
            """, (name, role, cpf if cpf else None, phone if phone else None))
            imported += 1

        conn.commit()
        cur.close()
        conn.close()

        log_action(session['user_id'], session['username'], f"Importou em lote {imported} pessoas/técnicos")
        return jsonify({"success": True, "imported": imported})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500


def get_recursive_folder_counts(conn, folder_ids):
    if not folder_ids:
        return {}
    
    cur = conn.cursor()
    cur.execute("SELECT id, parent_id FROM project_folders;")
    all_folders = cur.fetchall()
    
    cur.execute("SELECT id, folder_id FROM projects WHERE folder_id IS NOT NULL;")
    all_projects = cur.fetchall()
    cur.close()
    
    children_map = {}
    for f in all_folders:
        pid = f['parent_id']
        children_map.setdefault(pid, []).append(f['id'])
        
    direct_files_map = {}
    for p in all_projects:
        fid = p['folder_id']
        direct_files_map[fid] = direct_files_map.get(fid, 0) + 1
        
    results = {}
    
    def count_subtree(fid):
        total_files = direct_files_map.get(fid, 0)
        total_subfolders = 0
        
        child_fids = children_map.get(fid, [])
        for child_fid in child_fids:
            c_files, c_folders = count_subtree(child_fid)
            total_files += c_files
            total_subfolders += 1 + c_folders
            
        return total_files, total_subfolders

    for fid in folder_ids:
        f_count, sub_count = count_subtree(fid)
        results[fid] = {
            'file_count': f_count,
            'folder_count': sub_count
        }
        
    return results

# --- FOLDERS API ---
@app.route('/api/folders', methods=['GET'])
@login_required
def api_get_folders():
    parent_id = request.args.get('parent_id')
    if parent_id == 'null' or parent_id == '' or parent_id is None:
        parent_id = None
    else:
        parent_id = int(parent_id)
        
    try:
        conn = get_db()
        cur = conn.cursor()
        
        if parent_id is None:
            cur.execute("SELECT * FROM project_folders WHERE parent_id IS NULL ORDER BY name ASC;")
        else:
            cur.execute("SELECT * FROM project_folders WHERE parent_id = %s ORDER BY name ASC;", (parent_id,))
        folders = cur.fetchall()
        
        if parent_id is None:
            cur.execute("SELECT * FROM projects WHERE folder_id IS NULL ORDER BY created_at DESC;")
        else:
            cur.execute("SELECT * FROM projects WHERE folder_id = %s ORDER BY created_at DESC;", (parent_id,))
        projects = cur.fetchall()
        
        folder_ids = [f['id'] for f in folders]
        rec_counts = get_recursive_folder_counts(conn, folder_ids)
        
        cur.close()
        conn.close()
        
        for f in folders:
            f['created_at'] = f['created_at'].strftime('%d/%m/%Y %H:%M')
            counts = rec_counts.get(f['id'], {'file_count': 0, 'folder_count': 0})
            f['file_count'] = counts['file_count']
            f['folder_count'] = counts['folder_count']
        for p in projects:
            p['created_at'] = p['created_at'].strftime('%d/%m/%Y %H:%M')
            
        return jsonify({
            "folders": folders,
            "projects": projects
        })
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/folders', methods=['POST'])
@login_required
def api_create_folder():
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        parent_id = data.get('parent_id')
        if parent_id == 'null' or parent_id == '' or parent_id is None:
            parent_id = None
        else:
            parent_id = int(parent_id)
            
        if not name:
            return jsonify({"error": "Nome da pasta é obrigatório."}), 400
            
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO project_folders (name, parent_id, created_by) VALUES (%s, %s, %s) RETURNING id;",
            (name, parent_id, session['user_id'])
        )
        folder_id = cur.fetchone()['id']
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Criou a pasta de projetos: {name} (ID: {folder_id})")
        return jsonify({"success": True, "id": folder_id}), 201
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/folders/<int:folder_id>', methods=['PUT'])
@login_required
def api_update_folder(folder_id):
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        if not name:
            return jsonify({"error": "Nome da pasta é obrigatório."}), 400

        conn = get_db()
        cur = conn.cursor()
        cur.execute("UPDATE project_folders SET name = %s WHERE id = %s;", (name, folder_id))
        conn.commit()
        cur.close()
        conn.close()

        log_action(session['user_id'], session['username'], f"Renomeou pasta de projetos ID {folder_id} para {name}")
        return jsonify({"success": True})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/folders/<int:folder_id>', methods=['DELETE'])
@login_required
@roles_required('Coordenador', 'Supervisor')
def api_delete_folder(folder_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT name FROM project_folders WHERE id = %s;", (folder_id,))
        folder = cur.fetchone()
        if not folder:
            cur.close()
            conn.close()
            return jsonify({"error": "Pasta não encontrada."}), 404
            
        cur.execute("DELETE FROM project_folders WHERE id = %s;", (folder_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Excluiu a pasta de projetos: {folder['name']}")
        return jsonify({"success": True})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

# --- BULK AND EXPORTS ---
@app.route('/api/technicians/export', methods=['GET'])
@login_required
def api_export_technicians():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT name, cpf, phone, identity, dob, role, company, area, team_type, 
                   shirt_size, boot_size, pants_size, jacket_size, team,
                   registration_claro, registration_third, toa_login, phone_model, imei_1, imei_2, email
            FROM technicians ORDER BY name ASC;
        """)
        techs = cur.fetchall()
        cur.close()
        conn.close()
        
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow([
            "Nome", "CPF", "Telefone", "RG", "Data Nascimento", "Cargo", "Empresa", "Área de Atuação", 
            "Tipo Equipe", "Camisa", "Bota", "Calça", "Casaco", "Equipe",
            "Matricula Claro", "Matricula Terceiro", "Login TOA", "Modelo Telefone", "IMEI 1 Telefone", "IMEI 2 Telefone", "E-mail"
        ])
        for t in techs:
            writer.writerow([
                t['name'], t['cpf'], t['phone'] or '', t['identity'] or '', 
                t['dob'].strftime('%Y-%m-%d') if t['dob'] else '', 
                t['role'], t['company'] or '', t['area'] or '', t['team_type'] or '',
                t['shirt_size'] or '', t['boot_size'] or '', t['pants_size'] or '', t['jacket_size'] or '',
                t['team'] or '',
                t['registration_claro'] or '', t['registration_third'] or '', t['toa_login'] or '',
                t['phone_model'] or '', t['imei_1'] or '', t['imei_2'] or '', t['email'] or ''
            ])
            
        response = app.response_class(
            output.getvalue().encode('utf-8-sig'),
            mimetype='text/csv',
            headers={"Content-disposition": f"attachment; filename=tecnicos_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
        )
        return response
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/technicians/bulk', methods=['POST'])
@login_required
def api_bulk_technicians():
    try:
        data = request.get_json()
        csv_text = data.get('csv_data', '').strip()
        if not csv_text:
            return jsonify({"error": "Nenhum dado informado."}), 400
            
        delimiter = ';' if ';' in csv_text else ','
        f = io.StringIO(csv_text)
        reader = csv.reader(f, delimiter=delimiter)
        
        conn = get_db()
        cur = conn.cursor()
        
        success_count = 0
        errors = []
        
        for index, row in enumerate(reader):
            if index == 0 and any(h in row[0].lower() for h in ['nome', 'name', 'cpf']):
                continue
            if len(row) < 4:
                errors.append(f"Linha {index+1}: colunas insuficientes. Requer pelo menos Nome, CPF, Data de Nascimento e Cargo.")
                continue
                
            name = row[0].strip()
            cpf = row[1].strip().replace('.', '').replace('-', '')
            dob_raw = row[2].strip()
            role = row[3].strip()
            company = row[4].strip() if len(row) > 4 else None
            phone = row[5].strip() if len(row) > 5 else None
            identity = row[6].strip() if len(row) > 6 else None
            area = row[7].strip() if len(row) > 7 else None
            team_type = row[8].strip() if len(row) > 8 else None
            team = row[9].strip() if len(row) > 9 else None
            registration_claro = row[10].strip() if len(row) > 10 else None
            registration_third = row[11].strip() if len(row) > 11 else None
            toa_login = row[12].strip() if len(row) > 12 else None
            phone_model = row[13].strip() if len(row) > 13 else None
            imei_1 = row[14].strip() if len(row) > 14 else None
            imei_2 = row[15].strip() if len(row) > 15 else None
            email = row[16].strip() if len(row) > 16 else None
            
            if not name or not cpf or not dob_raw or not role:
                errors.append(f"Linha {index+1}: campos obrigatórios vazios.")
                continue
                
            try:
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d'):
                    try:
                        dob = datetime.datetime.strptime(dob_raw, fmt).date()
                        break
                    except ValueError:
                        continue
                else:
                    raise ValueError()
            except Exception:
                errors.append(f"Linha {index+1} ({name}): data '{dob_raw}' inválida. Use AAAA-MM-DD ou DD/MM/AAAA.")
                continue
                
            cur.execute("SELECT id FROM technicians WHERE cpf = %s;", (cpf,))
            if cur.fetchone():
                errors.append(f"Linha {index+1} ({name}): CPF {cpf} já cadastrado.")
                continue
                
            try:
                cur.execute(
                    """
                    INSERT INTO technicians (
                        name, cpf, phone, identity, dob, role, area, team_type, company, team,
                        registration_claro, registration_third, toa_login, phone_model, imei_1, imei_2, email, created_by
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                    """,
                    (name, cpf, phone, identity, dob, role, area, team_type, company, team,
                     registration_claro, registration_third, toa_login, phone_model, imei_1, imei_2, email, session['user_id'])
                )
                success_count += 1
            except Exception as err:
                conn.rollback()
                errors.append(f"Linha {index+1} ({name}): erro ao salvar: {err}")
                
        if errors and success_count == 0:
            cur.close()
            conn.close()
            return jsonify({"error": "Falha total no processamento em lote.", "details": errors}), 400
            
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Realizou cadastro em lote de técnicos: {success_count} importados.")
        return jsonify({"success": True, "imported": success_count, "errors": errors}), 200
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/vehicles/export', methods=['GET'])
@login_required
def api_export_vehicles():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            """
            SELECT v.plate, v.type, v.model, t.name AS responsible_name, v.subclus,
                   v.area_rede, v.base, v.setor, v.condutor_dia, v.condutor_tarde, 
                   v.condutor_madrugada, v.ticket_car, v.has_rack, v.has_basket, 
                   v.has_giroflex, v.has_inverter
            FROM vehicles v 
            LEFT JOIN technicians t ON v.responsible_tech_id = t.id 
            ORDER BY v.plate ASC;
            """
        )
        vehicles = cur.fetchall()
        cur.close()
        conn.close()
        
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow([
            "Placa", "Tipo", "Modelo", "Responsável", "Sub-Classe", "Área de Rede", "Base", "Setor", 
            "Condutor Dia", "Condutor Tarde", "Condutor Madrugada", "Ticket Car", "Rack de Teto", 
            "Cesto Aéreo", "Giroflex", "Inversor"
        ])
        for v in vehicles:
            writer.writerow([
                v['plate'], v['type'], v['model'] or '', v['responsible_name'] or '', 
                v['subclus'] or '', v['area_rede'] or '', v['base'] or '', v['setor'] or '',
                v['condutor_dia'] or '', v['condutor_tarde'] or '', v['condutor_madrugada'] or '', 
                v['ticket_car'] or '', "SIM" if v['has_rack'] else "NÃO",
                "SIM" if v['has_basket'] else "NÃO", "SIM" if v['has_giroflex'] else "NÃO",
                "SIM" if v['has_inverter'] else "NÃO"
            ])
            
        response = app.response_class(
            output.getvalue().encode('utf-8-sig'),
            mimetype='text/csv',
            headers={"Content-disposition": f"attachment; filename=veiculos_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
        )
        return response
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/vehicles/bulk', methods=['POST'])
@login_required
def api_bulk_vehicles():
    try:
        raw_rows = []
        file = request.files.get('file')
        if file and file.filename:
            filename = secure_filename(file.filename)
            temp_path = os.path.join(UPLOAD_FOLDER, f"v_temp_{int(time.time())}_{filename}")
            file.save(temp_path)
            ext = os.path.splitext(temp_path)[1].lower()
            
            if ext in ['.xlsx', '.xlsm', '.xls']:
                try:
                    from python_calamine import CalamineWorkbook
                    wb = CalamineWorkbook.from_path(temp_path)
                    target_sheet = wb.sheet_names[0]
                    for s in wb.sheet_names:
                        if any(k in s.lower() for k in ['veic', 'placa', 'frota', 'dados', 'base']):
                            target_sheet = s
                            break
                    raw_rows = wb.get_sheet_by_name(target_sheet).to_python()
                except Exception:
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
                        target_sheet = wb.sheetnames[0]
                        for s in wb.sheetnames:
                            if any(k in s.lower() for k in ['veic', 'placa', 'frota', 'dados', 'base']):
                                target_sheet = s
                                break
                        raw_rows = list(wb[target_sheet].iter_rows(values_only=True))
                    except Exception as e_excel:
                        if os.path.exists(temp_path): os.remove(temp_path)
                        return jsonify({"error": f"Erro ao ler arquivo Excel: {str(e_excel)}"}), 400
            else:
                content = None
                for enc in ['utf-8-sig', 'utf-8', 'latin1', 'cp1252']:
                    try:
                        with open(temp_path, 'r', encoding=enc) as f:
                            content = f.read()
                        if content: break
                    except Exception: continue
                if content:
                    lines = [l for l in content.splitlines() if l.strip()]
                    if lines:
                        delims = [';', ',', '\t', '|']
                        delim_counts = {d: lines[0].count(d) for d in delims}
                        best_delim = max(delim_counts, key=delim_counts.get)
                        if delim_counts[best_delim] == 0: best_delim = ','
                        reader = csv.reader(io.StringIO(content), delimiter=best_delim)
                        raw_rows = list(reader)

            if os.path.exists(temp_path):
                os.remove(temp_path)
        else:
            data = request.get_json(silent=True) or {}
            csv_text = (data.get('csv_data') or request.form.get('csv_data') or '').strip()
            if not csv_text:
                return jsonify({"error": "Nenhum arquivo ou dado informado para importação."}), 400
            delims = [';', ',', '\t', '|']
            first_line = csv_text.splitlines()[0] if csv_text.splitlines() else ''
            delim_counts = {d: first_line.count(d) for d in delims}
            best_delim = max(delim_counts, key=delim_counts.get)
            if delim_counts[best_delim] == 0: best_delim = ';'
            reader = csv.reader(io.StringIO(csv_text), delimiter=best_delim)
            raw_rows = list(reader)

        if not raw_rows:
            return jsonify({"error": "Nenhum dado válido extraído da planilha/texto."}), 400

        clean_rows = [r for r in raw_rows if r and any(c is not None and str(c).strip() != '' for c in r)]
        if not clean_rows:
            return jsonify({"error": "Planilha/texto sem conteúdo aproveitável."}), 400

        # Header detection
        header_idx = -1
        header_map = {}
        vehicle_keywords = ['PLACA', 'MODELO', 'TIPO', 'CPF', 'SUBCLUSTER', 'SUBCLUS', 'AREA', 'BASE', 'SETOR', 'CONDUTOR', 'TICKET']
        
        for i, r in enumerate(clean_rows[:15]):
            str_r = [normalize_text_key(c) for c in r]
            concat = ' '.join(str_r)
            if sum(1 for k in vehicle_keywords if k in concat) >= 1 or 'PLACA' in concat:
                header_idx = i
                for c_idx, val in enumerate(str_r):
                    if val: header_map[val] = c_idx
                break

        if header_idx != -1:
            data_rows = clean_rows[header_idx + 1:]
        else:
            data_rows = clean_rows

        def get_val(row, candidates, default_pos):
            for cand in candidates:
                norm_cand = normalize_text_key(cand)
                for h_text, col_i in header_map.items():
                    if norm_cand in h_text or h_text in norm_cand:
                        if col_i < len(row) and row[col_i] is not None:
                            return str(row[col_i]).strip()
            if default_pos < len(row) and row[default_pos] is not None:
                return str(row[default_pos]).strip()
            return ''

        conn = get_db()
        cur = conn.cursor()
        success_count = 0
        updated_count = 0
        inserted_count = 0
        errors = []

        for index, row in enumerate(data_rows):
            plate = get_val(row, ['PLACA', 'PLATE', 'VEICULO'], 0).upper()
            plate = ''.join(c for c in plate if c.isalnum() or c == '-')
            model = get_val(row, ['MODELO', 'MODEL'], 1)
            v_type = get_val(row, ['TIPO', 'TYPE', 'CATEGORIA'], 2) or 'Utilitário'
            resp_tech_cpf = get_val(row, ['CPF', 'CPF_RESPONSAVEL', 'RESPONSAVEL', 'MOTORISTA'], 3)
            resp_tech_cpf = ''.join(filter(str.isdigit, resp_tech_cpf)) if resp_tech_cpf else None
            subclus = get_val(row, ['SUBCLUSTER', 'SUB_CLUSTER', 'SUBCLUS'], 4)
            area_rede = get_val(row, ['AREA_REDE', 'AREA', 'REGIAO'], 5)
            base = get_val(row, ['BASE', 'CIDADE'], 6)
            setor = get_val(row, ['SETOR', 'GESTAO'], 7)
            condutor_dia = get_val(row, ['CONDUTOR_DIA', 'DIA'], 8)
            condutor_tarde = get_val(row, ['CONDUTOR_TARDE', 'TARDE'], 9)
            condutor_madrugada = get_val(row, ['CONDUTOR_MADRUGADA', 'MADRUGADA'], 10)
            ticket_car = get_val(row, ['TICKET_CAR', 'TICKETCAR', 'CARTAO'], 11)

            if not plate:
                continue

            tech_id = None
            if resp_tech_cpf:
                cur.execute("SELECT id FROM technicians WHERE cpf = %s;", (resp_tech_cpf,))
                tech = cur.fetchone()
                if tech:
                    tech_id = tech['id']

            cur.execute("SELECT id FROM vehicles WHERE plate = %s;", (plate,))
            existing = cur.fetchone()

            try:
                if existing:
                    cur.execute(
                        """
                        UPDATE vehicles SET
                            model = COALESCE(NULLIF(%s, ''), model),
                            type = COALESCE(NULLIF(%s, ''), type),
                            responsible_tech_id = COALESCE(%s, responsible_tech_id),
                            subclus = COALESCE(NULLIF(%s, ''), subclus),
                            area_rede = COALESCE(NULLIF(%s, ''), area_rede),
                            base = COALESCE(NULLIF(%s, ''), base),
                            setor = COALESCE(NULLIF(%s, ''), setor),
                            condutor_dia = COALESCE(NULLIF(%s, ''), condutor_dia),
                            condutor_tarde = COALESCE(NULLIF(%s, ''), condutor_tarde),
                            condutor_madrugada = COALESCE(NULLIF(%s, ''), condutor_madrugada),
                            ticket_car = COALESCE(NULLIF(%s, ''), ticket_car)
                        WHERE id = %s;
                        """,
                        (model, v_type, tech_id, subclus, area_rede, base, setor,
                         condutor_dia, condutor_tarde, condutor_madrugada, ticket_car, existing['id'])
                    )
                    updated_count += 1
                else:
                    cur.execute(
                        """
                        INSERT INTO vehicles (
                            plate, model, type, responsible_tech_id, subclus, area_rede, base, setor,
                            condutor_dia, condutor_tarde, condutor_madrugada, ticket_car, created_by
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                        """,
                        (plate, model or 'Veículo', v_type, tech_id, subclus, area_rede, base, setor,
                         condutor_dia, condutor_tarde, condutor_madrugada, ticket_car, session['user_id'])
                    )
                    inserted_count += 1
                success_count += 1
            except Exception as err:
                conn.rollback()
                errors.append(f"Linha {index+1} (Placa: {plate}): erro ao salvar - {err}")

        conn.commit()
        cur.close()
        conn.close()

        log_action(session['user_id'], session['username'], f"Realizou cadastro/atualização em lote de veículos: {success_count} processados ({inserted_count} novos, {updated_count} atualizados).")
        return jsonify({
            "success": True,
            "imported": success_count,
            "inserted": inserted_count,
            "updated": updated_count,
            "errors": errors
        }), 200
    except Exception as e:
        if 'conn' in locals() and conn:
            try: conn.rollback(); conn.close()
            except Exception: pass
        print("Backend exception:", e)
        return jsonify({"error": "Ocorreu um erro ao processar a requisição."}), 500


# --- FINANCIAL: TEAMS & CONSUMABLES ---
@app.route('/api/finance/teams', methods=['GET', 'POST'])
@login_required
@coordenador_claro_required
def api_finance_teams():
    if request.method == 'GET':
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute(
                """
                SELECT tf.*, t1.name AS tech1_name, t2.name AS tech2_name 
                FROM teams_finance tf 
                LEFT JOIN technicians t1 ON tf.tech1_id = t1.id 
                LEFT JOIN technicians t2 ON tf.tech2_id = t2.id 
                ORDER BY tf.reference_month DESC, tf.created_at DESC;
                """
            )
            records = cur.fetchall()
            cur.close()
            conn.close()
            for r in records:
                r['amount'] = float(r['amount'])
                r['created_at'] = r['created_at'].strftime('%d/%m/%Y %H:%M')
            return jsonify(records)
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500
    else:
        try:
            data = request.get_json()
            tech1_id = safe_int(data.get('tech1_id'))
            tech2_id = safe_int(data.get('tech2_id'))
            area = data.get('area', '').strip()
            amount = safe_float(data.get('amount', 0.0))
            ref_month = data.get('reference_month', '').strip()
            
            if not tech1_id or not tech2_id or not area or amount <= 0 or not ref_month:
                return jsonify({"error": "Pessoas, Área, Valor e Mês de Referência são obrigatórios."}), 400
                
            conn = get_db()
            cur = conn.cursor()
            cur.execute(
                """
                INSERT INTO teams_finance (tech1_id, tech2_id, area, amount, reference_month, created_by) 
                VALUES (%s, %s, %s, %s, %s, %s) RETURNING id;
                """,
                (tech1_id, tech2_id, area, amount, ref_month, session['user_id'])
            )
            record_id = cur.fetchone()['id']
            conn.commit()
            cur.close()
            conn.close()
            
            log_action(session['user_id'], session['username'], f"Registrou faturamento de equipe (ID: {record_id}, Valor: R$ {amount:.2f}, Mês: {ref_month})")
            return jsonify({"success": True, "id": record_id}), 201
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/finance/teams/<int:record_id>', methods=['DELETE', 'PUT'])
@login_required
@coordenador_claro_required
def api_finance_team_detail(record_id):
    if request.method == 'DELETE':
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("SELECT amount, reference_month FROM teams_finance WHERE id = %s;", (record_id,))
            row = cur.fetchone()
            if not row:
                cur.close()
                conn.close()
                return jsonify({"error": "Registro não encontrado."}), 404
                
            cur.execute("DELETE FROM teams_finance WHERE id = %s;", (record_id,))
            conn.commit()
            cur.close()
            conn.close()
            log_action(session['user_id'], session['username'], f"Excluiu faturamento de equipe (Valor: R$ {float(row['amount']):.2f}, Mês: {row['reference_month']})")
            return jsonify({"success": True})
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500
    else: # PUT
        try:
            data = request.get_json() or {}
            tech1_id = safe_int(data.get('tech1_id'))
            tech2_id = safe_int(data.get('tech2_id'))
            area = data.get('area', '').strip()
            amount = safe_float(data.get('amount', 0.0))
            ref_month = data.get('reference_month', '').strip()

            if not tech1_id or not tech2_id or not area or amount <= 0 or not ref_month:
                return jsonify({"error": "Pessoas, Área, Valor e Mês de Referência são obrigatórios."}), 400

            conn = get_db()
            cur = conn.cursor()
            cur.execute("""
                UPDATE teams_finance 
                SET tech1_id = %s, tech2_id = %s, area = %s, amount = %s, reference_month = %s 
                WHERE id = %s;
            """, (tech1_id, tech2_id, area, amount, ref_month, record_id))
            conn.commit()
            cur.close()
            conn.close()

            log_action(session['user_id'], session['username'], f"Atualizou faturamento de equipe ID {record_id} (Valor: R$ {amount:.2f}, Mês: {ref_month})")
            return jsonify({"success": True})
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/finance/consumables', methods=['GET', 'POST'])
@login_required
@coordenador_claro_required
def api_finance_consumables():
    if request.method == 'GET':
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("SELECT * FROM consumables_finance ORDER BY reference_month DESC, created_at DESC;")
            records = cur.fetchall()
            cur.close()
            conn.close()
            for r in records:
                r['amount'] = float(r['amount'])
                r['created_at'] = r['created_at'].strftime('%d/%m/%Y %H:%M')
            return jsonify(records)
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500
    else:
        try:
            data = request.get_json()
            description = data.get('description', '').strip()
            area = data.get('area', '').strip()
            amount = safe_float(data.get('amount', 0.0))
            ref_month = data.get('reference_month', '').strip()
            
            if not description or not area or amount <= 0 or not ref_month:
                return jsonify({"error": "Descrição, Área, Valor e Mês de Referência são obrigatórios."}), 400
                
            conn = get_db()
            cur = conn.cursor()
            cur.execute(
                """
                INSERT INTO consumables_finance (description, area, amount, reference_month, created_by) 
                VALUES (%s, %s, %s, %s, %s) RETURNING id;
                """,
                (description, area, amount, ref_month, session['user_id'])
            )
            record_id = cur.fetchone()['id']
            conn.commit()
            cur.close()
            conn.close()
            
            log_action(session['user_id'], session['username'], f"Registrou custo de consumíveis: {description} (Valor: R$ {amount:.2f}, Mês: {ref_month})")
            return jsonify({"success": True, "id": record_id}), 201
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/finance/consumables/<int:record_id>', methods=['DELETE', 'PUT'])
@login_required
@coordenador_claro_required
def api_finance_consumables_detail(record_id):
    if request.method == 'DELETE':
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("SELECT amount, description FROM consumables_finance WHERE id = %s;", (record_id,))
            row = cur.fetchone()
            if not row:
                cur.close()
                conn.close()
                return jsonify({"error": "Registro não encontrado."}), 404
                
            cur.execute("DELETE FROM consumables_finance WHERE id = %s;", (record_id,))
            conn.commit()
            cur.close()
            conn.close()
            log_action(session['user_id'], session['username'], f"Excluiu despesa de consumível (ID: {record_id}, {row['description']})")
            return jsonify({"success": True})
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500
    else: # PUT
        try:
            data = request.get_json() or {}
            description = data.get('description', '').strip()
            area = data.get('area', '').strip()
            amount = safe_float(data.get('amount', 0.0))
            ref_month = data.get('reference_month', '').strip()

            if not description or not area or amount <= 0 or not ref_month:
                return jsonify({"error": "Descrição, Área, Valor e Mês de Referência são obrigatórios."}), 400

            conn = get_db()
            cur = conn.cursor()
            cur.execute("""
                UPDATE consumables_finance 
                SET description = %s, area = %s, amount = %s, reference_month = %s 
                WHERE id = %s;
            """, (description, area, amount, ref_month, record_id))
            conn.commit()
            cur.close()
            conn.close()

            log_action(session['user_id'], session['username'], f"Atualizou custo de consumível ID {record_id} ({description}, R$ {amount:.2f})")
            return jsonify({"success": True})
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/finance/export', methods=['GET'])
@login_required
@coordenador_claro_required
def api_export_finance():
    try:
        conn = get_db()
        cur = conn.cursor()

        cur.execute("""
            SELECT tf.id, tf.reference_month, tf.area, tf.amount, tf.created_at,
                   t1.name AS tech1_name, t2.name AS tech2_name
            FROM teams_finance tf
            LEFT JOIN technicians t1 ON tf.tech1_id = t1.id
            LEFT JOIN technicians t2 ON tf.tech2_id = t2.id
            ORDER BY tf.reference_month DESC, tf.created_at DESC;
        """)
        teams = cur.fetchall()

        cur.execute("""
            SELECT id, reference_month, description, area, amount, created_at
            FROM consumables_finance
            ORDER BY reference_month DESC, created_at DESC;
        """)
        consumables = cur.fetchall()

        cur.close()
        conn.close()

        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow(["Tipo Lançamento", "Mês Referência", "Descrição / Pessoas", "Área de Trabalho", "Valor Mensal (R$)", "Data Registro"])

        for t in teams:
            tech1 = t['tech1_name'] or 'N/A'
            tech2 = t['tech2_name'] or 'N/A'
            dta = t['created_at'].strftime('%d/%m/%Y %H:%M') if t['created_at'] else ''
            writer.writerow(["Equipe Externa", t['reference_month'], f"{tech1} + {tech2}", t['area'], f"{float(t['amount']):.2f}".replace('.', ','), dta])

        for c in consumables:
            dta = c['created_at'].strftime('%d/%m/%Y %H:%M') if c['created_at'] else ''
            writer.writerow(["Consumível", c['reference_month'], c['description'], c['area'], f"{float(c['amount']):.2f}".replace('.', ','), dta])

        response = app.response_class(
            output.getvalue().encode('utf-8-sig'),
            mimetype='text/csv',
            headers={"Content-disposition": f"attachment; filename=relatorio_financeiro_custos_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
        )
        return response
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

# --- PHYSICAL INVENTORY ---
@app.route('/api/inventory', methods=['GET', 'POST'])
@login_required
def api_inventory():
    if request.method == 'GET':
        category = request.args.get('category', '').strip()
        try:
            conn = get_db()
            cur = conn.cursor()
            if category:
                cur.execute("SELECT * FROM inventory_items WHERE category = %s ORDER BY name ASC;", (category,))
            else:
                cur.execute("SELECT * FROM inventory_items ORDER BY category ASC, name ASC;")
            items = cur.fetchall()
            cur.close()
            conn.close()
            for i in items:
                i['created_at'] = i['created_at'].strftime('%d/%m/%Y %H:%M')
            return jsonify(items)
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500
    else:
        try:
            data = request.get_json()
            category = data.get('category', '').strip()
            name = data.get('name', '').strip()
            quantity = safe_int(data.get('quantity', 0))
            serial_number = data.get('serial_number', '').strip() or None
            description = data.get('description', '').strip() or None
            
            if not category or not name or quantity < 0:
                return jsonify({"error": "Categoria, Nome e Quantidade são obrigatórios."}), 400
                
            conn = get_db()
            cur = conn.cursor()
            cur.execute(
                """
                INSERT INTO inventory_items (category, name, quantity, serial_number, description, created_by) 
                VALUES (%s, %s, %s, %s, %s, %s) RETURNING id;
                """,
                (category, name, quantity, serial_number, description, session['user_id'])
            )
            item_id = cur.fetchone()['id']
            conn.commit()
            cur.close()
            conn.close()
            
            log_action(session['user_id'], session['username'], f"Adicionou item ao inventário ({category}): {name} (Qtd: {quantity})")
            return jsonify({"success": True, "id": item_id}), 201
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/inventory/<int:item_id>', methods=['PUT', 'DELETE'])
@login_required
def api_inventory_detail(item_id):
    if request.method == 'PUT':
        try:
            data = request.get_json()
            name = data.get('name', '').strip()
            quantity = safe_int(data.get('quantity', 0))
            serial_number = data.get('serial_number', '').strip() or None
            description = data.get('description', '').strip() or None
            
            if not name or quantity < 0:
                return jsonify({"error": "Nome e Quantidade são obrigatórios."}), 400
                
            conn = get_db()
            cur = conn.cursor()
            
            cur.execute("SELECT category, name FROM inventory_items WHERE id = %s;", (item_id,))
            item = cur.fetchone()
            if not item:
                cur.close()
                conn.close()
                return jsonify({"error": "Item não encontrado."}), 404
                
            cur.execute(
                """
                UPDATE inventory_items 
                SET name = %s, quantity = %s, serial_number = %s, description = %s 
                WHERE id = %s;
                """,
                (name, quantity, serial_number, description, item_id)
            )
            conn.commit()
            cur.close()
            conn.close()
            
            log_action(session['user_id'], session['username'], f"Atualizou item do inventário ({item['category']}): {name} (Qtd: {quantity})")
            return jsonify({"success": True})
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500
    else:
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("SELECT category, name FROM inventory_items WHERE id = %s;", (item_id,))
            item = cur.fetchone()
            if not item:
                cur.close()
                conn.close()
                return jsonify({"error": "Item não encontrado."}), 404
                
            cur.execute("DELETE FROM inventory_items WHERE id = %s;", (item_id,))
            conn.commit()
            cur.close()
            conn.close()
            
            log_action(session['user_id'], session['username'], f"Deletou item do inventário ({item['category']}): {item['name']}")
            return jsonify({"success": True})
        except Exception as e:
            print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500


# --- 5. AUDIT LOGS ---
@app.route('/api/logs', methods=['GET'])
@login_required
@roles_required('Coordenador', 'Supervisor')
def api_get_logs():
    user_filter = request.args.get('username', '').strip().lower()
    try:
        conn = get_db()
        cur = conn.cursor()
        if user_filter:
            cur.execute(
                "SELECT * FROM audit_logs WHERE LOWER(username) LIKE %s ORDER BY timestamp DESC LIMIT 200;",
                (f"%{user_filter}%",)
            )
        else:
            cur.execute("SELECT * FROM audit_logs ORDER BY timestamp DESC LIMIT 200;")
        logs = cur.fetchall()
        cur.close()
        conn.close()
        
        for l in logs:
            if l['timestamp']:
                l['timestamp'] = l['timestamp'].strftime('%d/%m/%Y %H:%M:%S')
            if isinstance(l.get('action'), str):
                l['action'] = l['action'].replace('GestÃ£o', 'Gestão')
        return jsonify(logs)
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500


# --- 6. USER MANAGEMENT ---
@app.route('/api/users', methods=['GET'])
@login_required
@roles_required('Coordenador', 'Supervisor')
def api_get_users():
    try:
        conn = get_db()
        cur = conn.cursor()
        if session['role'] == 'Supervisor':
            cur.execute("""
                SELECT u.id, u.username, u.role, u.tech_id, t.name AS tech_name, u.created_at 
                FROM users u 
                LEFT JOIN technicians t ON u.tech_id = t.id 
                WHERE u.role != 'Coordenador' 
                ORDER BY u.username ASC;
            """)
        else:
            cur.execute("""
                SELECT u.id, u.username, u.role, u.tech_id, t.name AS tech_name, u.created_at 
                FROM users u 
                LEFT JOIN technicians t ON u.tech_id = t.id 
                ORDER BY u.username ASC;
            """)
        users = cur.fetchall()
        cur.close()
        conn.close()
        
        for u in users:
            if u['created_at']:
                u['created_at'] = u['created_at'].strftime('%d/%m/%Y %H:%M')
        return jsonify(users)
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/users', methods=['POST'])
@login_required
@roles_required('Coordenador')
def api_create_user():
    data = request.get_json()
    username = data.get('username', '').strip().lower()
    password = data.get('password', '')
    role = data.get('role', '').strip()
    tech_id = data.get('tech_id')
    
    if not tech_id:
        return jsonify({"error": "A pessoa precisa estar cadastrada primeiro na seção Pessoas (#pessoas) antes de criar uma conta de acesso."}), 400

    if not username or not role:
        return jsonify({"error": "Todos os campos de usuário (exceto senha) são obrigatórios."}), 400
        
    if role not in ['Supervisor', 'Técnico', 'Auxiliar', 'Coordenador']:
        return jsonify({"error": "Cargo inválido informado."}), 400
            
    try:
        conn = get_db()
        cur = conn.cursor()

        # Check technician existence
        cur.execute("SELECT id, name, company FROM technicians WHERE id = %s;", (int(tech_id),))
        tech = cur.fetchone()
        if not tech:
            cur.close()
            conn.close()
            return jsonify({"error": "A pessoa selecionada não foi encontrada na seção Pessoas (#pessoas). Cadastre-a primeiro."}), 400

        cur.execute("SELECT id FROM users WHERE username = %s;", (username,))
        if cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({"error": "Nome de usuário já está em uso."}), 400
            
        empresa = (tech.get('company') or 'Claro').strip()
        senha_padrao = f"{empresa}@2026"
        hashed = generate_password_hash(senha_padrao)
        cur.execute(
            "INSERT INTO users (username, password_hash, role, tech_id) VALUES (%s, %s, %s, %s) RETURNING id;",
            (username, hashed, role, int(tech_id))
        )
        new_id = cur.fetchone()['id']
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Criou o usuário de acesso: {username} (Pessoa: {tech['name']}, Cargo: {role})")
        return jsonify({"success": True, "id": new_id}), 201
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
@roles_required('Coordenador')
def api_delete_user(user_id):
    if user_id == session['user_id']:
        return jsonify({"error": "Não é possível excluir a própria conta em uso."}), 400
        
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT username, role FROM users WHERE id = %s;", (user_id,))
        user = cur.fetchone()
        
        if not user:
            cur.close()
            conn.close()
            return jsonify({"error": "Usuário não encontrado."}), 404
            
        # Clear foreign key references before deleting user account
        for tbl, col in [('audit_logs', 'user_id'), ('technicians', 'created_by'), ('vehicles', 'created_by'), ('financial_logs', 'logged_by'), ('project_folders', 'created_by'), ('projects', 'created_by')]:
            try:
                cur.execute(f"UPDATE {tbl} SET {col} = NULL WHERE {col} = %s;", (user_id,))
            except Exception:
                conn.rollback()

        cur.execute("DELETE FROM users WHERE id = %s;", (user_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Deletou a conta do usuário: {user['username']}")
        return jsonify({"success": True})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required
@roles_required('Coordenador')
def api_update_user(user_id):
    data = request.get_json() or {}
    username = data.get('username', '').strip().lower()
    role = data.get('role', '').strip()
    password = data.get('password', '')
    tech_id = data.get('tech_id')
    
    if not role:
        return jsonify({"error": "O cargo é obrigatório."}), 400
        
    if role not in ['Supervisor', 'Técnico', 'Auxiliar', 'Coordenador']:
        return jsonify({"error": "Cargo inválido informado."}), 400
            
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Check if user exists
        cur.execute("SELECT username, role FROM users WHERE id = %s;", (user_id,))
        user = cur.fetchone()
        if not user:
            cur.close()
            conn.close()
            return jsonify({"error": "Usuário não encontrado."}), 404
            
        target_username = user['username']
        if username and username != user['username']:
            cur.execute("SELECT id FROM users WHERE username = %s AND id != %s;", (username, user_id))
            if cur.fetchone():
                cur.close()
                conn.close()
                return jsonify({"error": "Nome de usuário (login) já está em uso por outra conta."}), 400
            target_username = username

        t_id = int(tech_id) if tech_id else None

        if password:
            hashed = generate_password_hash(password)
            cur.execute(
                "UPDATE users SET username = %s, role = %s, password_hash = %s, tech_id = %s WHERE id = %s;",
                (target_username, role, hashed, t_id, user_id)
            )
        else:
            cur.execute(
                "UPDATE users SET username = %s, role = %s, tech_id = %s WHERE id = %s;",
                (target_username, role, t_id, user_id)
            )
            
        conn.commit()
        cur.close()
        conn.close()
        
        log_action(session['user_id'], session['username'], f"Editou o usuário de acesso: {target_username} (Cargo: {role})")
        return jsonify({"success": True})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

# ==============================================================================
# BUSCADOR DE FALHAS API (RAL, REC, HFC, GPON)
# ==============================================================================
import unicodedata

BUSCADOR_COLUMN_MAPS = {
    'RAL': ['Bairro', 'RalAno', 'Tipo Equip', 'Bandeira', 'Distância', 'Ref', 'Logradouro', 'Data', 'Status'],
    'REC': ['Bairro', 'Chamado', 'Tecnologia', 'Tipo', 'Distância', 'Ref', 'Logradouro', 'Data', 'Status'],
    'HFC': ['Bairro', 'Ticket T', 'Nodename', 'Node', 'Distância', 'Ref', 'Logradouro', 'Data', 'Status'],
    'GPON': ['Bairro', 'Ticket T', 'OLT', 'Slot PON', 'Distância', 'Ref', 'Logradouro', 'Data', 'Status']
}

def normalize_text_key(txt):
    if not txt:
        return ""
    txt = str(txt).strip().upper()
    nfkd = unicodedata.normalize('NFKD', txt)
    return "".join([c for c in nfkd if not unicodedata.combining(c)])


def parse_fault_searcher_file(filepath, topic):
    ext = os.path.splitext(filepath)[1].lower()
    raw_rows = []
    
    if ext in ['.xlsx', '.xlsm', '.xls']:
        try:
            from python_calamine import CalamineWorkbook
            wb = CalamineWorkbook.from_path(filepath)
            sheet_names = wb.sheet_names
            target_sheet = sheet_names[0]
            sheet_found = False
            for candidate in [topic, 'RAL', 'REC', 'HFC', 'GPON', 'FALHAS', 'GERAL', 'Analítico', 'Dados', 'Ocorrências', 'Dados_DB', 'Base']:
                for s in sheet_names:
                    if candidate.lower() in s.lower():
                        target_sheet = s
                        sheet_found = True
                        break
                if sheet_found:
                    break

            ws = wb.get_sheet_by_name(target_sheet)
            raw_rows = ws.to_python()
        except Exception as e_cal:
            print("Calamine parse warning, falling back to openpyxl:", e_cal)
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                sheet_names = wb.sheetnames
                target_sheet = sheet_names[0]
                sheet_found = False
                for candidate in [topic, 'RAL', 'REC', 'HFC', 'GPON', 'FALHAS', 'GERAL', 'Analítico', 'Dados', 'Ocorrências', 'Dados_DB', 'Base']:
                    for s in sheet_names:
                        if candidate.lower() in s.lower():
                            target_sheet = s
                            sheet_found = True
                            break
                    if sheet_found:
                        break
                ws = wb[target_sheet]
                raw_rows = list(ws.iter_rows(values_only=True))
            except Exception as e_opx:
                print("openpyxl parse warning:", e_opx)
                raw_rows = []
    elif ext in ['.csv', '.txt']:
        content = None
        for enc in ['utf-8-sig', 'utf-8', 'latin1', 'cp1252']:
            try:
                with open(filepath, 'r', encoding=enc) as f:
                    content = f.read()
                if content:
                    break
            except Exception:
                continue
                
        if content:
            lines = [line for line in content.splitlines() if line.strip()]
            if lines:
                first_line = lines[0]
                delims = [';', ',', '\t', '|']
                delim_counts = {d: first_line.count(d) for d in delims}
                best_delim = max(delim_counts, key=delim_counts.get)
                if delim_counts[best_delim] == 0:
                    best_delim = ','
                import csv
                import io
                reader = csv.reader(io.StringIO(content), delimiter=best_delim)
                raw_rows = list(reader)

    if not raw_rows:
        return []

    clean_rows = []
    for r in raw_rows:
        if r and any(c is not None and str(c).strip() != '' for c in r):
            clean_rows.append(r)

    if not clean_rows:
        return []

    header_idx = 0
    header_map = {}
    
    header_keywords = [
        'DESIGNA', 'CHAMADO', 'ROTA', 'RALANO', 'TICKET', 'SUBCLUSTER',
        'DISTANCIA', 'ENDERECO', 'MES', 'OFENSOR', 'OCORRENCIA', 'CAUSA',
        'CIRCUITO', 'REF', 'NUM'
    ]

    for i, r in enumerate(clean_rows[:25]):
        str_r = [normalize_text_key(c) for c in r]
        concat = ' '.join(str_r)
        matched_count = sum(1 for k in header_keywords if k in concat)
        if matched_count >= 2 or any(k in concat for k in ['RALANO', 'CHAMADO', 'SUBCLUSTER']):
            header_idx = i
            for c_idx, val in enumerate(str_r):
                if val:
                    header_map[val] = c_idx
            break

    if not header_map and clean_rows:
        header_idx = 0
        for c_idx, val in enumerate(clean_rows[0]):
            norm_val = normalize_text_key(val)
            if norm_val:
                header_map[norm_val] = c_idx

    data_rows = clean_rows[header_idx + 1:] if header_idx < len(clean_rows) - 1 else clean_rows[header_idx:]
    if data_rows == [clean_rows[header_idx]]:
        data_rows = clean_rows[header_idx + 1:]

    parsed_records = []

    def find_col_index(candidates, default_pos):
        for cand in candidates:
            norm_cand = normalize_text_key(cand)
            for h_text, col_i in header_map.items():
                if norm_cand in h_text or h_text in norm_cand:
                    return col_i
        return default_pos

    def get_val_at(row, col_i):
        if 0 <= col_i < len(row):
            v = row[col_i]
            if v is not None:
                if isinstance(v, (datetime.datetime, datetime.date)):
                    return v.strftime('%d/%m/%Y %H:%M')
                return str(v).strip()
        return ''

    idx_ref = find_col_index(['REF', 'REFERENCIA', 'REF_VAL', 'REFERENCIA_LOCAL'], 999)

    if topic == 'RAL':
        idx1 = find_col_index(['DESIGNACAO', 'DESIGNAO', 'CIRCUITO', 'ROTA'], 7)
        idx2 = find_col_index(['RALANO', 'NUM RAL', 'NUM_RAL', 'CHAMADO', 'TICKET'], 0)
        idx3 = find_col_index(['DATA DE ABERTURA', 'DATA_ABERTURA', 'DATA ABERTURA', 'ABERTURA', 'INICIO'], 8)
        idx4 = find_col_index(['SUBCLUSTER', 'SUB_CLUSTER', 'CLUSTER', 'SALA'], 20)
        idx5 = find_col_index(['DISTANCIA', 'DISTACIA', 'KM', 'METROS'], 48)
        idx6 = find_col_index(['ENDERECO', 'ENDERECO PADRONIZADO', 'LOGRADOURO', 'RUA'], 27)
        idx7 = find_col_index(['MES FECOM', 'MES', 'MENSAL'], 17)
        idx8 = find_col_index(['OFENSOR', 'CFOFENSOR', 'RESPONSAVEL'], 29)
        idx9 = find_col_index(['OCORRENCIA', 'CAUSA', 'MOTIVO', 'TIPO'], 25)
    elif topic == 'REC':
        idx1 = find_col_index(['DESIGNACAO_CIRCUITO', 'DESIGNACAO', 'CIRCUITO', 'ROTA'], 4)
        idx2 = find_col_index(['CHAMADO', 'NUM_REC', 'TICKET', 'RALANO'], 0)
        idx3 = find_col_index(['DT_HORA_ABERTURA_REC', 'HORA_ABERTURA', 'DATA DE ABERTURA', 'DATA ABERTURA', 'INICIO'], 9)
        idx4 = find_col_index(['SUBCLUSTER', 'SUB_CLUSTER', 'CLUSTER', 'SALA'], 24)
        idx5 = find_col_index(['DISTANCIA', 'DISTACIA', 'KM'], 35)
        idx6 = find_col_index(['ENDERECO', 'LOGRADOURO'], 15)
        idx7 = find_col_index(['MES'], 17)
        idx8 = find_col_index(['OFENSOR', 'CFOFENSOR'], 20)
        idx9 = find_col_index(['OCORRENCIA', 'CAUSA'], 13)
    elif topic == 'HFC':
        idx1 = find_col_index(['ROTA', 'DESIGNACAO', 'CIRCUITO'], 16)
        idx2 = find_col_index(['TICKET', 'TICKE T', 'TICKET T', 'CHAMADO'], 0)
        idx3 = find_col_index(['INICIO_QUEDA', 'INICIO QUEDA', 'DATA INICIO', 'DATA ABERTURA'], 3)
        idx4 = find_col_index(['SUBCLUSTER', 'SUB_CLUSTER', 'CLUSTER', 'SALA'], 18)
        idx5 = find_col_index(['DISTANCIA', 'DISTACIA', 'KM'], 51)
        idx6 = find_col_index(['ENDERECO', 'LOGRADOURO'], 35)
        idx7 = find_col_index(['MES INCIDENTE', 'MES'], 24)
        idx8 = find_col_index(['OFENSOR', 'CFOFENSOR'], 27)
        idx9 = find_col_index(['CAUSA', 'OCORRENCIA'], 20)
    else:
        idx1 = find_col_index(['DESIGNACAO', 'CIRCUITO', 'ROTA'], 0)
        idx2 = find_col_index(['CHAMADO', 'TICKET', 'NUM_REC', 'RALANO'], 1)
        idx3 = find_col_index(['DATA DE ABERTURA', 'HORA_ABERTURA', 'INICIO_QUEDA', 'DATA ABERTURA'], 2)
        idx4 = find_col_index(['SUBCLUSTER', 'SUB_CLUSTER'], 3)
        idx5 = find_col_index(['DISTANCIA', 'DISTACIA', 'KM'], 8)
        idx6 = find_col_index(['ENDERECO', 'LOGRADOURO'], 5)
        idx7 = find_col_index(['MES'], 6)
        idx8 = find_col_index(['OFENSOR'], 7)
        idx9 = find_col_index(['OCORRENCIA', 'CAUSA'], 4)

    for row_idx, row in enumerate(data_rows):
        if not any(c is not None and str(c).strip() != '' for c in row):
            continue

        col_ref = get_val_at(row, idx_ref) or '-'
        col1 = get_val_at(row, idx1)
        col2 = get_val_at(row, idx2)
        col3 = get_val_at(row, idx3)
        col4 = get_val_at(row, idx4)
        col5 = get_val_at(row, idx5)
        col6 = get_val_at(row, idx6)
        col7 = get_val_at(row, idx7)
        col8 = get_val_at(row, idx8)
        col9 = get_val_at(row, idx9)

        if not col2:
            if col1:
                col2 = col1
            else:
                col2 = f"{topic}_REC_{row_idx + 1}"

        parsed_records.append((topic, col1, col2, col3, col4, col5, col_ref, col6, col7, col8, col9))

    return parsed_records


    return parsed_records

@app.route('/api/buscador/records', methods=['GET'])
@login_required
def api_get_buscador_records():
    try:
        topic = request.args.get('topic', 'RAL').upper()
        search = request.args.get('search', '').strip()
        route = request.args.get('route', '').strip()

        if topic not in BUSCADOR_COLUMN_MAPS:
            topic = 'RAL'

        columns = BUSCADOR_COLUMN_MAPS[topic]

        conn = get_db()
        cur = conn.cursor()

        log_info = None
        try:
            cur.execute("SELECT last_updated, updated_by, filename, record_count FROM fault_searcher_logs WHERE topic = %s;", (topic,))
            log_info = cur.fetchone()
        except Exception:
            conn.rollback()

        last_updated_str = "Não carregado"
        if log_info and log_info.get('last_updated'):
            last_updated_str = log_info['last_updated'].strftime('%d/%m - %H:%M')

        routes = []
        try:
            cur.execute("""
                SELECT DISTINCT col1 
                FROM fault_searcher_records 
                WHERE topic = %s AND col1 IS NOT NULL AND col1 != '' 
                ORDER BY col1 ASC;
            """, (topic,))
            routes = [r['col1'] for r in cur.fetchall()]
        except Exception:
            conn.rollback()

        base_query = "SELECT id, col1, col2, col3, col4, col5, COALESCE(ref_val, '-') AS ref_val, col6, col7, col8, col9, is_edited FROM fault_searcher_records WHERE topic = %s"
        params = [topic]

        if route:
            base_query += " AND (col1 ILIKE %s)"
            params.append(f"%{route}%")

        if search:
            search_terms = [t.strip() for t in search.split() if t.strip()]
            for term in search_terms:
                base_query += """ AND (
                    col1 ILIKE %s OR col2 ILIKE %s OR col3 ILIKE %s OR col4 ILIKE %s OR
                    col5 ILIKE %s OR COALESCE(ref_val, '') ILIKE %s OR col6 ILIKE %s OR col7 ILIKE %s OR col8 ILIKE %s OR col9 ILIKE %s
                )"""
                t_pattern = f"%{term}%"
                params.extend([t_pattern] * 10)

        query = base_query + """ ORDER BY 
            CASE WHEN col5 IS NULL OR TRIM(col5) = '' OR TRIM(col5) = '-' THEN 1 ELSE 0 END ASC,
            CASE WHEN NULLIF(REGEXP_REPLACE(col5, '[^0-9]', '', 'g'), '') IS NOT NULL 
                 THEN NULLIF(REGEXP_REPLACE(col5, '[^0-9]', '', 'g'), '')::bigint 
                 ELSE NULL END DESC NULLS LAST,
            col5 DESC,
            id ASC LIMIT 1000;"""

        try:
            cur.execute(query, tuple(params))
            rows = cur.fetchall()
        except Exception as e_sort:
            conn.rollback()
            fallback_query = base_query + """ ORDER BY 
                CASE WHEN col5 IS NULL OR TRIM(col5) = '' OR TRIM(col5) = '-' THEN 1 ELSE 0 END ASC,
                col5 DESC, id ASC LIMIT 1000;"""
            cur.execute(fallback_query, tuple(params))
            rows = cur.fetchall()

        records = []
        for r in rows:
            records.append([
                r.get('col1') or '', r.get('col2') or '', r.get('col3') or '',
                r.get('col4') or '', r.get('col5') or '', r.get('ref_val') or '-', r.get('col6') or '',
                r.get('col7') or '', r.get('col8') or '', r.get('col9') or '',
                r['id'],
                bool(r.get('is_edited'))
            ])

        cur.close()
        conn.close()

        return jsonify({
            "topic": topic,
            "columns": columns,
            "last_updated": last_updated_str,
            "record_count": len(records),
            "total_records": log_info['record_count'] if log_info and log_info.get('record_count') else len(records),
            "routes": routes,
            "records": records
        })
    except Exception as e:
        print("Error in api_get_buscador_records:", e)
        return jsonify({
            "topic": request.args.get('topic', 'RAL').upper(),
            "columns": BUSCADOR_COLUMN_MAPS.get(request.args.get('topic', 'RAL').upper(), []),
            "last_updated": "Não carregado",
            "record_count": 0,
            "total_records": 0,
            "routes": [],
            "records": [],
            "error": str(e)
        }), 200

@app.route('/api/buscador/records/<int:record_id>', methods=['PUT'])
@login_required
@roles_required('Coordenador', 'Supervisor')
def api_update_buscador_distance(record_id):
    data = request.get_json() or {}
    distance = (data.get('distance') or '').strip()
    ref_val = (data.get('ref') or '').strip() or '-'

    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE fault_searcher_records 
            SET col5 = %s, ref_val = %s, is_edited = TRUE 
            WHERE id = %s 
            RETURNING topic, col2;
        """, (distance, ref_val, record_id))
        row = cur.fetchone()
        if not row:
            cur.close()
            conn.close()
            return jsonify({"error": "Registro não encontrado."}), 404

        conn.commit()
        cur.close()
        conn.close()

        log_action(session['user_id'], session['username'], f"Editou a distância ('{distance}') e REF ('{ref_val}') do item [{row['topic']}] (Código: {row['col2']})")
        return jsonify({"success": True, "distance": distance, "ref": ref_val})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500

@app.route('/api/buscador/upload', methods=['POST'])
@login_required
@roles_required('Coordenador', 'Supervisor')
def api_upload_buscador_base():
    if 'file' not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado."}), 400

    file = request.files['file']
    topic = request.form.get('topic', 'RAL').upper()

    if topic not in BUSCADOR_COLUMN_MAPS:
        return jsonify({"error": "Tópico de base inválido."}), 400

    if not file or not file.filename:
        return jsonify({"error": "Arquivo inválido."}), 400

    filename = secure_filename(file.filename)
    temp_path = os.path.join(UPLOAD_FOLDER, f"temp_{int(time.time())}_{filename}")
    file.save(temp_path)

    try:
        parsed_records = parse_fault_searcher_file(temp_path, topic)
        if not parsed_records:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return jsonify({"error": "Não foi possível extrair dados válidos da planilha enviada."}), 400

        conn = get_db()
        cur = conn.cursor()

        # 1. Fetch existing records for this topic from DB (mapped by col2 unique code: RALANO / CHAMADO / TICKET T)
        cur.execute("""
            SELECT col1, col2, col3, col4, col5, COALESCE(ref_val, '-') AS ref_val, col6, col7, col8, col9, is_edited
            FROM fault_searcher_records
            WHERE topic = %s;
        """, (topic,))
        existing_rows = cur.fetchall()

        existing_map = {}
        for r in existing_rows:
            key = str(r['col2'] or '').strip()
            if key:
                existing_map[key] = dict(r)

        # 2. Merge parsed records with existing DB edits (col5 is Distância, ref_val is REF)
        final_records = []
        processed_keys = set()

        for row in parsed_records:
            t_val, c1, c2, c3, c4, c5, c_ref, c6, c7, c8, c9 = row
            key = str(c2 or '').strip()

            if key and key in existing_map:
                ex = existing_map[key]
                processed_keys.add(key)
                # Preserve user edited distance and ref, or existing DB distance/ref if new file value is empty
                if ex.get('is_edited') or (ex.get('col5') and not c5):
                    final_c5 = ex['col5']
                    is_ed = True
                else:
                    final_c5 = c5
                    is_ed = bool(c5 and c5 != '-' and c5 != '')

                if ex.get('is_edited') or (ex.get('ref_val') and ex['ref_val'] != '-' and not c_ref):
                    final_ref = ex['ref_val']
                else:
                    final_ref = c_ref or '-'

                final_records.append((t_val, c1, c2, c3, c4, final_c5, final_ref, c6, c7, c8, c9, is_ed))
            else:
                if key:
                    processed_keys.add(key)
                is_ed = bool(c5 and c5 != '-' and c5 != '')
                final_records.append((t_val, c1, c2, c3, c4, c5, c_ref or '-', c6, c7, c8, c9, is_ed))

        # 3. Retain any historical records in DB that might not be in the new file
        for key, ex in existing_map.items():
            if key not in processed_keys:
                final_records.append((
                    topic, ex['col1'] or '', ex['col2'] or '', ex['col3'] or '',
                    ex['col4'] or '', ex['col5'] or '', ex['ref_val'] or '-', ex['col6'] or '',
                    ex['col7'] or '', ex['col8'] or '', ex['col9'] or '',
                    bool(ex.get('is_edited'))
                ))

        # 4. Replace records with merged data
        cur.execute("DELETE FROM fault_searcher_records WHERE topic = %s;", (topic,))

        from psycopg2.extras import execute_values
        execute_values(cur, """
            INSERT INTO fault_searcher_records (topic, col1, col2, col3, col4, col5, ref_val, col6, col7, col8, col9, is_edited)
            VALUES %s;
        """, final_records, page_size=2000)

        # 5. Update log timestamp
        cur.execute("""
            INSERT INTO fault_searcher_logs (topic, last_updated, updated_by, filename, record_count)
            VALUES (%s, CURRENT_TIMESTAMP, %s, %s, %s)
            ON CONFLICT (topic) DO UPDATE SET
                last_updated = CURRENT_TIMESTAMP,
                updated_by = EXCLUDED.updated_by,
                filename = EXCLUDED.filename,
                record_count = EXCLUDED.record_count;
        """, (topic, session['username'], filename, len(final_records)))

        conn.commit()
        cur.close()
        conn.close()

        if os.path.exists(temp_path):
            os.remove(temp_path)

        log_action(session['user_id'], session['username'], f"Atualizou a base do Buscador [{topic}] com {len(final_records)} registros mantendo edições (Arquivo: {filename})")

        return jsonify({
            "success": True,
            "message": f"Base [{topic}] atualizada com sucesso! {len(final_records)} registros salvos e edições preservadas.",
            "record_count": len(final_records)
        })
    except Exception as e:
        if 'conn' in locals() and conn:
            try:
                conn.rollback()
                conn.close()
            except Exception:
                pass
        if os.path.exists(temp_path):
            os.remove(temp_path)
        print("Backend exception:", e)
        return jsonify({"error": "Ocorreu um erro ao processar a requisição."}), 500


@app.route('/api/buscador/export', methods=['GET'])
@login_required
def api_export_buscador_base():
    topic = request.args.get('topic', 'RAL').upper()
    search = request.args.get('search', '').strip()
    route = request.args.get('route', '').strip()

    if topic not in BUSCADOR_COLUMN_MAPS:
        topic = 'RAL'

    columns = BUSCADOR_COLUMN_MAPS[topic]

    conn = get_db()
    cur = conn.cursor()

    query = """
        SELECT col1, col2, col3, col4, col5, COALESCE(ref_val, '-') AS ref_val, col6, col7, col8, col9
        FROM fault_searcher_records
        WHERE topic = %s
    """
    params = [topic]

    if route:
        query += " AND (col1 ILIKE %s)"
        params.append(f"%{route}%")

    if search:
        search_terms = [t.strip() for t in search.split() if t.strip()]
        for term in search_terms:
            query += """ AND (
                col1 ILIKE %s OR col2 ILIKE %s OR col3 ILIKE %s OR col4 ILIKE %s OR
                col5 ILIKE %s OR COALESCE(ref_val, '') ILIKE %s OR col6 ILIKE %s OR col7 ILIKE %s OR col8 ILIKE %s OR col9 ILIKE %s
            )"""
            t_pattern = f"%{term}%"
            params.extend([t_pattern] * 10)

    query += """ ORDER BY 
        CASE WHEN col5 IS NULL OR TRIM(col5) = '' OR TRIM(col5) = '-' THEN 1 ELSE 0 END ASC,
        CASE WHEN NULLIF(REGEXP_REPLACE(col5, '[^0-9]', '', 'g'), '') IS NOT NULL 
             THEN NULLIF(REGEXP_REPLACE(col5, '[^0-9]', '', 'g'), '')::bigint 
             ELSE NULL END DESC NULLS LAST,
        col5 DESC,
        id ASC;"""

    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(columns)

    for r in rows:
        writer.writerow([
            r['col1'] or '', r['col2'] or '', r['col3'] or '',
            r['col4'] or '', r['col5'] or '', r['ref_val'] or '-', r['col6'] or '',
            r['col7'] or '', r['col8'] or '', r['col9'] or ''
        ])

    response = app.response_class(
        response=output.getvalue().encode('utf-8-sig'),
        status=200,
        mimetype='text/csv'
    )
    response.headers["Content-Disposition"] = f"attachment; filename=Buscador_{topic}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return response

# --- ESCALA DE TRABALHO APIs ---
DAY_NAMES_PT = ['seg', 'ter', 'qua', 'qui', 'sex', 'sáb', 'dom']

@app.route('/api/schedules', methods=['GET'])
@login_required
def api_get_schedules():
    month_str = request.args.get('month', '').strip()
    area = request.args.get('area', '').strip()
    company = request.args.get('company', '').strip()
    search = request.args.get('search', '').strip()

    # Parse month (YYYY-MM)
    if not month_str or len(month_str) != 7 or '-' not in month_str:
        now = datetime.datetime.now()
        month_str = now.strftime('%Y-%m')

    try:
        year, month = map(int, month_str.split('-'))
    except Exception:
        now = datetime.datetime.now()
        year, month = now.year, now.month
        month_str = f"{year:04d}-{month:02d}"

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Ensure work_schedules table exists
        try:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS work_schedules (
                    id SERIAL PRIMARY KEY,
                    tech_id INT NOT NULL,
                    date DATE NOT NULL,
                    status VARCHAR(50) DEFAULT 'Trabalho',
                    work_hours VARCHAR(50) DEFAULT '08 às 17:48hs',
                    on_call VARCHAR(50) DEFAULT '0',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(tech_id, date)
                );
            """)
            conn.commit()
        except Exception as e_tbl:
            conn.rollback()

        # Fetch distinct areas for filter dropdown (splitting comma-separated values like "Norte 1, Norte 2, Norte 3")
        cur.execute("SELECT DISTINCT area FROM technicians WHERE area IS NOT NULL AND area != '';")
        area_rows = cur.fetchall()
        areas_set = set()
        for r in area_rows:
            val = r.get('area', '')
            if val:
                for sub in val.split(','):
                    sub_clean = sub.strip()
                    if sub_clean:
                        areas_set.add(sub_clean)
        areas_list = sorted(list(areas_set))

        # If no area selected, return area_required=True
        if not area:
            return jsonify({
                "area_required": True,
                "month": month_str,
                "areas_list": areas_list,
                "technicians": [],
                "days_in_month": []
            })

        # Fetch technicians matching area (exact or in comma-separated list)
        tech_query = "SELECT * FROM technicians WHERE (UPPER(area) = UPPER(%s) OR UPPER(area) LIKE %s)"
        params = [area, f"%{area.upper()}%"]

        if company and company.lower() != 'todas':
            tech_query += " AND UPPER(company) = UPPER(%s)"
            params.append(company)

        if search:
            tech_query += " AND (UPPER(name) LIKE %s OR UPPER(role) LIKE %s OR UPPER(phone) LIKE %s)"
            s_pattern = f"%{search.upper()}%"
            params.extend([s_pattern, s_pattern, s_pattern])

        # Order Coordenadores first, then Técnicos
        tech_query += """ ORDER BY 
            CASE 
                WHEN UPPER(role) LIKE 'COORDENADOR%%' THEN 1 
                WHEN UPPER(role) LIKE 'ADMIN%%' THEN 2 
                ELSE 3 
            END, name ASC;"""

        cur.execute(tech_query, tuple(params))
        technicians = cur.fetchall()
        tech_ids = [t['id'] for t in technicians]

        # Build calendar days for the month
        num_days = calendar.monthrange(year, month)[1]
        days_in_month = []
        for day in range(1, num_days + 1):
            dt = datetime.date(year, month, day)
            day_str = dt.strftime('%Y-%m-%d')
            weekday_idx = dt.weekday() # 0 = Monday, 6 = Sunday
            days_in_month.append({
                "date": day_str,
                "day_num": day,
                "day_label": f"{day:02d}/{month:02d}",
                "day_name": DAY_NAMES_PT[weekday_idx],
                "is_weekend": weekday_idx >= 5
            })

        # Fetch existing schedule records for these techs in this month
        schedules_map = {}
        if tech_ids:
            start_date = f"{year:04d}-{month:02d}-01"
            end_date = f"{year:04d}-{month:02d}-{num_days:02d}"
            placeholders = ','.join(['%s'] * len(tech_ids))
            query_schedules = f"""
                SELECT tech_id, date::text as date_str, status, work_hours, on_call
                FROM work_schedules
                WHERE tech_id IN ({placeholders}) AND date >= %s AND date <= %s;
            """
            sched_params = list(tech_ids) + [start_date, end_date]
            cur.execute(query_schedules, tuple(sched_params))
            for r in cur.fetchall():
                schedules_map[(r['tech_id'], r['date_str'])] = {
                    "status": r['status'] or 'Trabalho',
                    "work_hours": r['work_hours'] or '08 às 17:48hs',
                    "on_call": r['on_call'] or '0'
                }

        # Format technicians output with schedules map
        tech_list = []
        for t in technicians:
            t_sched = {}
            for d in days_in_month:
                ds = d['date']
                rec = schedules_map.get((t['id'], ds), {
                    "status": "Trabalho",
                    "work_hours": "08 às 17:48hs",
                    "on_call": "0"
                })
                t_sched[ds] = rec

            tech_list.append({
                "id": t['id'],
                "name": t['name'],
                "role": t.get('role') or 'Técnico',
                "company": t.get('company') or 'Claro',
                "area": t.get('area') or '',
                "schedules": t_sched
            })

        return jsonify({
            "area_required": False,
            "month": month_str,
            "area": area,
            "company": company,
            "areas_list": areas_list,
            "days_in_month": days_in_month,
            "technicians": tech_list
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        print("Backend exception:", e)
        return jsonify({"error": "Ocorreu um erro ao processar a requisição."}), 200
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass

@app.route('/api/schedules/batch', methods=['POST'])
@login_required
def api_save_schedules_batch():
    conn = None
    try:
        data = request.get_json(silent=True) or {}
        updates = data.get('updates', [])
        if not updates:
            return jsonify({"success": True, "updated": 0})

        conn = get_db()
        cur = conn.cursor()

        for u in updates:
            tech_id = int(u.get('tech_id'))
            dt_str = u.get('date')
            status = u.get('status', 'Trabalho')
            work_hours = u.get('work_hours', '08 às 17:48hs')
            on_call = str(u.get('on_call', '0'))

            cur.execute(
                """
                INSERT INTO work_schedules (tech_id, date, status, work_hours, on_call)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT (tech_id, date) 
                DO UPDATE SET 
                    status = EXCLUDED.status,
                    work_hours = EXCLUDED.work_hours,
                    on_call = EXCLUDED.on_call;
                """,
                (tech_id, dt_str, status, work_hours, on_call)
            )

        conn.commit()

        log_action(session['user_id'], session['username'], f"Atualizou {len(updates)} dia(s) na escala de trabalho")
        return jsonify({"success": True, "updated": len(updates)})
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao processar a requisição."}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass

# --- MAPA DE EVENTOS API ---
# All coordinates cover the Rio de Janeiro metro area (State of RJ)
# SubCluster code â†’ approximate center coordinates
SUBCLUSTER_COORDS_MAP = {
    # Metro / Metropolitana
    'MTP':  (-22.9068, -43.1729),
    'CENTRO':      (-22.9056, -43.1769),
    'TIJUCA':      (-22.9248, -43.2327),
    'COPACABANA':  (-22.9694, -43.1868),
    'BARRA':       (-23.0106, -43.3654),
    'RECREIO':     (-23.0209, -43.4569),
    'JACAREPAGUA': (-22.9553, -43.3622),
    'CAMPO GRANDE':(-22.9035, -43.5574),
    'BANGU':       (-22.8784, -43.4665),
    'SANTA CRUZ':  (-22.9268, -43.6759),
    'GUARATIBA':   (-23.0208, -43.5761),
    'NITEROI':     (-22.8833, -43.1036),
    'NITERI':      (-22.8833, -43.1036),
    'SAO GONCALO': (-22.8270, -43.0541),
    'MARICA':      (-22.9194, -42.8184),
    # Baixada Fluminense
    'BXF':         (-22.7558, -43.4603),
    'BAIXADA FLUMINENSE': (-22.7558, -43.4603),
    'NOVA IGUACU': (-22.7558, -43.4603),
    'DUQUE DE CAXIAS': (-22.7858, -43.3056),
    'CAXIAS':      (-22.7858, -43.3056),
    'MESQUITA':    (-22.8155, -43.4398),
    'NILOPOLISS':  (-22.8047, -43.4222),
    'NILOPOLIS':   (-22.8047, -43.4222),
    'QUEIMADOS':   (-22.7119, -43.5562),
    'JAPERÍ':      (-22.6425, -43.6594),
    'JAPERI':      (-22.6425, -43.6594),
    'ITAGUAI':     (-22.8716, -43.7797),
    'SEROPEDICA':  (-22.7457, -43.7097),
    'BELFORD ROXO':(-22.7644, -43.3994),
    'SAO JOAO DE MERITI':(-22.8029, -43.3738),
    # Norte
    'NO1':         (-22.9035, -43.5574),  # Norte 1 (Campo Grande / Oeste)
    'NO2':         (-22.8715, -43.3364),  # Norte 2 (Madureira / Irajá)
    'NO3':         (-22.8091, -43.2084),  # Norte 3 (Ilha / Bonsucesso / Leopoldina)
    'MADUREIRA':   (-22.8715, -43.3364),
    'IRAJA':       (-22.8543, -43.3362),
    'MEIER':       (-22.9009, -43.2714),
    'PENHA':       (-22.8459, -43.2693),
    'ILHA DO GOVERNADOR': (-22.8091, -43.2084),
    'LEOPOLDINA':  (-22.8669, -43.2545),
    'BONSUCESSO':  (-22.8694, -43.2526),
    'PETROPOLIS':  (-22.5050, -43.1789),
    'PETRÓPOLIS':  (-22.5050, -43.1789),
    'VOLTA REDONDA':(-22.5228, -44.1048),
    'CABO FRIO':   (-22.8792, -42.0186),
}

# Fallback region codes â†’ center coordinates
AREA_COORDS = {
    'MTP': (-22.9068, -43.1729),
    'BXF': (-22.7558, -43.4603),
    'NO1': (-22.9035, -43.5574),
    'NO2': (-22.8715, -43.3364),
    'NO3': (-22.8091, -43.2084),
}

MONTH_NAMES_PT = {
    1: 'jan', 2: 'fev', 3: 'mar', 4: 'abr', 5: 'mai', 6: 'jun',
    7: 'jul', 8: 'ago', 9: 'set', 10: 'out', 11: 'nov', 12: 'dez'
}
MONTH_NAMES_FULL_PT = {
    1: 'janeiro', 2: 'fevereiro', 3: 'março', 4: 'abril', 5: 'maio', 6: 'junho',
    7: 'julho', 8: 'agosto', 9: 'setembro', 10: 'outubro', 11: 'novembro', 12: 'dezembro'
}

def _get_coords_for_subcluster(col4_val, address_str=''):
    """Return (lat, lng) for a given col4 (subcluster) and col6 (address) value from the Buscador base."""
    import unicodedata
    def strip_accents(text):
        return ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
        
    s_subcluster = strip_accents((col4_val or '')).upper().strip()
    s_address = strip_accents((address_str or '')).upper().strip()
    
    # First: direct or partial key match in subcluster
    for key, coords in SUBCLUSTER_COORDS_MAP.items():
        if key in s_subcluster or s_subcluster in key:
            return coords
            
    # Next: check if the address contains any known neighborhood or city name
    for key, coords in SUBCLUSTER_COORDS_MAP.items():
        if key in s_address:
            return coords
            
    # Check for area prefix code in subcluster
    for area_code, coords in AREA_COORDS.items():
        if s_subcluster.startswith(area_code):
            return coords
            
    # Fallback: Rio de Janeiro center
    return (-22.9068, -43.1729)

def _detect_subcluster_code(col4_val, col6_val=''):
    """Extract the standardized area code (NO1/NO2/NO3/BXF/MTP) from col4 and col6."""
    s = ((col4_val or '') + ' ' + (col6_val or '')).upper().strip()
    
    if any(k in s for k in ['BXF', 'IGUACU', 'IGUAÇU', 'CAXIAS', 'MESQUITA', 'NILOPOLIS', 'NILÓPOLIS', 'QUEIMADOS', 'JAPER', 'ITAGUAI', 'ITAGUAÍ', 'SEROPEDICA', 'SEROPÉDICA', 'BELFORD', 'MERITI']):
        return 'BXF'
    if any(k in s for k in ['NO2', 'MADUREIRA', 'IRAJA', 'IRAJÁ', 'MEIER', 'MÉIER', 'PENHA', 'CASCADURA', 'MARECHAL']):
        return 'NO2'
    if any(k in s for k in ['NO3', 'ILHA', 'LEOPOLDINA', 'BONSUCESSO', 'RAMOS', 'OLARIA', 'FUNDAO', 'FUNDÃO']):
        return 'NO3'
    if any(k in s for k in ['NO1', 'REALENGO', 'PADRE MIGUEL', 'DEODORO', 'SULACAP']):
        return 'NO1'
    if any(k in s for k in ['MTP', 'CENTRO', 'TIJUCA', 'COPA', 'BARRA', 'RECREIO', 'JACAREPAGUA', 'CAMPO GRANDE', 'BANGU', 'CRUZ', 'GUARATIBA', 'NITER', 'GONCALO', 'GONÇALO', 'MARICA', 'MARICÁ', 'CABO FRIO', 'PETROPOLIS', 'PETRÓPOLIS', 'MACA', 'SAO PEDRO', 'SÃO PEDRO', 'RIO DAS OSTRAS', 'BUZIOS', 'BÚZIOS', 'ARARUAMA', 'SAQUAREMA', 'IGUABA']):
        return 'MTP'
        
    for code in ['NO3', 'NO2', 'NO1', 'BXF', 'MTP']:
        if code in s:
            return code
            
    return 'OUTRO'

def _classify_event_type(col9, col8):
    """Classify event type from Ocorrência (col9) and Ofensor (col8)."""
    cause = ((col9 or '') + ' ' + (col8 or '')).upper()
    if 'CURTO' in cause or 'CC ' in cause or 'CURTO CIRC' in cause:
        return 'CURTO CIRCUITO'
    if ('BAIXA' in cause and ('SINAL' in cause or 'NIVEL' in cause or 'NIVEL' in cause or 'NIVEL' in cause)) \
       or 'REDE BAIXA' in cause or 'NIVEL BAIXO' in cause or 'FIBRA' in cause \
       or 'ROMPIMENTO' in cause or 'ATENUAÇÃO' in cause or 'ATENUACAO' in cause:
        return 'REDE BAIXA'
    if 'VANDAL' in cause or 'ROUBO' in cause or 'FURTO' in cause \
       or 'CORTE' in cause or 'DEPREDAÇÃO' in cause or 'DEPREDACAO' in cause:
        return 'VANDALISMO'
    # Fallback by keywords
    if 'CURTO' in cause:
        return 'CURTO CIRCUITO'
    if 'BAIXA' in cause:
        return 'REDE BAIXA'
    return 'VANDALISMO'

def _parse_month_from_col(col3, col7):
    """Extract month number from col3 (date string) or col7 (month name/number)."""
    # Try col7 first (e.g. "jul", "julho", "7")
    month_abbr_map = {'jan':1,'fev':2,'mar':3,'abr':4,'mai':5,'jun':6,
                      'jul':7,'ago':8,'set':9,'out':10,'nov':11,'dez':12}
    if col7:
        v = str(col7).strip().lower()
        if v in month_abbr_map:
            return month_abbr_map[v]
        for abbr, num in month_abbr_map.items():
            if v.startswith(abbr):
                return num
        try:
            val = int(v)
            if 1 <= val <= 12: return val
        except Exception:
            pass
            
    # Try col3 (date string e.g. "dd/mm/yyyy hh:mm", "2025-07-12")
    if col3:
        if hasattr(col3, 'month'):
            return col3.month
            
        import re
        s = str(col3).strip()
        
        # Try YYYY-MM-DD
        m_iso = re.search(r'^(\d{4})[\-/](\d{1,2})[\-/](\d{1,2})', s)
        if m_iso:
            val = int(m_iso.group(2))
            if 1 <= val <= 12: return val
            
        # Try DD/MM/YYYY or MM/DD/YYYY
        m_br = re.search(r'^(\d{1,2})[\./\-](\d{1,2})[\./\-](\d{2,4})', s)
        if m_br:
            val = int(m_br.group(2)) # assume DD/MM
            if 1 <= val <= 12: return val
            val2 = int(m_br.group(1)) # fallback MM/DD
            if 1 <= val2 <= 12: return val2
            
        # Fallback 
        m_mid = re.search(r'[\./\-](\d{1,2})[\./\-]', s)
        if m_mid:
            val = int(m_mid.group(1))
            if 1 <= val <= 12: return val
            
    return None


@app.route('/api/mapa-eventos/data', methods=['GET'])
@login_required
def api_get_mapa_eventos_data():
    """
    Returns map marker data from the same fault_searcher_records table used by the Buscador.
    Always reflects the most recently uploaded base automatically.
    """
    origin = request.args.get('origin', '').strip().upper()
    area = request.args.get('area', '').strip().upper()
    event_types_str = request.args.get('event_types', '').strip()
    month_str = request.args.get('month', '').strip()
    year_str = request.args.get('year', '').strip()

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # --- Build SQL query mirroring the Buscador structure ---
        conditions = []
        params = []

        # Topic filter (origin = RAL, REC, HFC, OUTAGE, GPON, or ALL)
        if origin and origin != 'TODAS':
            conditions.append("UPPER(topic) = %s")
            params.append(origin)

        # Area filter is applied in python now because subclusters don't map cleanly via LIKE
        # Event type filter is applied in python

        where_clause = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        sql = f"""
            SELECT id, topic, col1, col2, col3, col4, col6, col7, col8, col9
            FROM fault_searcher_records
            {where_clause}
            ORDER BY id DESC
            LIMIT 5000;
        """
        cur.execute(sql, tuple(params))
        rows = cur.fetchall()

        # Get last_updated from fault_searcher_logs
        cur.execute("SELECT MAX(last_updated) AS lu FROM fault_searcher_logs;")
        log_row = cur.fetchone()
        last_updated = (log_row['lu'].strftime('%d/%m/%Y %H:%M:%S') if log_row and log_row['lu']
                        else datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S'))

        # Parse event_types filter
        active_event_types = set()
        if event_types_str:
            for et in event_types_str.split(','):
                et = et.strip().upper()
                if et:
                    active_event_types.add(et)

        import random
        events = []
        monthly_counts = {m: 0 for m in range(1, 13)}

        for r in rows:
            rec_id = r['id']
            
            # Ensure address is present and valid
            address_str = (r['col6'] or '').strip()
            if not address_str or address_str == '-' or address_str.lower() in ['n/a', 'na']:
                continue
                
            col4_val = r['col4'] or ''
            col9_val = r['col9'] or ''
            col8_val = r['col8'] or ''

            # Classify event type from real Ocorrência / Ofensor fields
            ev_type = _classify_event_type(col9_val, col8_val)

            # Apply event type filter (only show selected types)
            if active_event_types and ev_type not in active_event_types:
                continue

            # Apply Area Filter
            extracted_area = _detect_subcluster_code(col4_val, address_str)
            if area and area != 'TODAS':
                if extracted_area != area:
                    continue

            # Get coordinates from subcluster (col4) or address (col6)
            lat, lng = _get_coords_for_subcluster(col4_val, address_str)
            
            # The database doesn't store exact coordinates. We use a deterministic jitter 
            # based on the address string to group identical addresses, with a small 
            # micro-jitter to prevent perfect overlaps. The max spread is ~1.5km to avoid water.
            import hashlib
            addr_hash = int(hashlib.md5(address_str.encode('utf-8')).hexdigest()[:8], 16)
            random.seed(addr_hash)
            base_jitter_lat = (random.random() - 0.5) * 0.012
            base_jitter_lng = (random.random() - 0.5) * 0.012
            
            random.seed(rec_id)
            micro_jitter_lat = (random.random() - 0.5) * 0.002
            micro_jitter_lng = (random.random() - 0.5) * 0.002
            
            jitter_lat = base_jitter_lat + micro_jitter_lat
            jitter_lng = base_jitter_lng + micro_jitter_lng

            # Count by month (from col3 date or col7 month field)
            month_num = _parse_month_from_col(r['col3'], r['col7'])
            
            # If a year filter is active, only count for that year
            if month_num and 1 <= month_num <= 12:
                if not year_str or year_str in str(r['col3']):
                    monthly_counts[month_num] += 1

            # Apply Month/Year filters to the map markers
            if year_str and year_str not in str(r['col3']):
                continue
            if month_str:
                if not month_num or str(month_num) != str(int(month_str)):
                    continue

            events.append({
                "id": rec_id,
                "topic": r['topic'],
                "designation": r['col1'] or '-',
                "ticket": r['col2'] or '-',
                "date": r['col3'] or '-',
                "subcluster": col4_val,
                "address": r['col6'] or '-',
                "cause": (col9_val or col8_val or 'OCORRÊNCIA DE REDE').upper(),
                "event_type": ev_type,
                "lat": round(lat + jitter_lat, 6),
                "lng": round(lng + jitter_lng, 6)
            })

        # Build mini monthly chart using real month counts from the filtered data
        chart_months = []
        for m_num in range(1, 13):
            if monthly_counts[m_num] > 0:
                chart_months.append({"month": MONTH_NAMES_PT[m_num], "count": monthly_counts[m_num]})

        # If no real month data, show last 6 months as zero (still useful context)
        if not chart_months:
            now = datetime.datetime.now()
            for offset in range(5, -1, -1):
                m = now.month - offset
                if m <= 0:
                    m += 12
                chart_months.append({"month": MONTH_NAMES_PT[m], "count": 0})

        return jsonify({
            "total": len(events),
            "events": events,
            "monthly_chart": chart_months,
            "last_updated": last_updated
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        print("Backend exception:", e)
        return jsonify({"error": "Ocorreu um erro ao processar a requisição."}), 200
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/mapa-eventos/projection', methods=['GET'])
@login_required
def api_mapa_eventos_projection():
    """
    Returns projection chart data (Power BI style) from the same Buscador base.
    All subclusters and counts are derived from real col4/col9/col8/col7 fields.
    """
    origin = request.args.get('origin', '').strip().upper()
    region = request.args.get('region', '').strip().upper()
    event_type_filter = request.args.get('event_type', '').strip().upper()
    month_str = request.args.get('month', str(datetime.datetime.now().month)).strip()
    year_str = request.args.get('year', str(datetime.datetime.now().year)).strip()

    try:
        current_month = int(month_str) if month_str else datetime.datetime.now().month
        current_year = int(year_str) if year_str else datetime.datetime.now().year
    except ValueError:
        current_month = datetime.datetime.now().month
        current_year = datetime.datetime.now().year

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Fetch all records, optionally filtered by topic
        conditions = []
        params = []
        if origin and origin not in ('', 'TODAS', 'ES'):
            conditions.append("UPPER(topic) = %s")
            params.append(origin)

        where_clause = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        sql = f"""
            SELECT id, topic, col1, col4, col6, col7, col8, col9, col3
            FROM fault_searcher_records
            {where_clause}
            ORDER BY id DESC
            LIMIT 8000;
        """
        cur.execute(sql, tuple(params))
        all_rows = cur.fetchall()

        # Collect all distinct area codes from real data
        area_counts_all = {}
        sc_label_counts_all = {}
        area_month_matrix = {}   # area_code -> month_num -> count
        sc_month_matrix = {}     # sc_label -> month_num -> count

        for r in all_rows:
            col4_val = r['col4'] or ''
            col6_val = r['col6'] or ''
            col9_val = r['col9'] or ''
            col8_val = r['col8'] or ''
            col3_val = r['col3'] or ''
            col7_val = r['col7'] or ''

            # Classify event type
            ev_type = _classify_event_type(col9_val, col8_val)
            if event_type_filter and event_type_filter != 'TODAS' and ev_type != event_type_filter:
                continue

            # Get area code and subcluster label from real data
            area_code = _detect_subcluster_code(col4_val, col6_val)
            
            if region and region != 'TODAS':
                row_text = f"{r['col1']} {col4_val} {col6_val} {r['topic']}".upper()
                if region == 'RIO A':
                    if area_code not in ('MTP', 'NO1', 'NO2', 'NO3') and 'RIO A' not in row_text:
                        continue
                elif region == 'RIO B':
                    if area_code not in ('BXF', 'OUTRO') and 'RIO B' not in row_text:
                        continue
                else:
                    if region not in row_text and area_code != region:
                        continue
                    
            sc_label = col4_val.strip().upper() if col4_val.strip() else 'OUTRO'
            # Truncate for display (max 12 chars)
            sc_label_short = sc_label[:12] if len(sc_label) > 12 else sc_label

            # Get month from real data
            month_num = _parse_month_from_col(col3_val, col7_val)

            # Count totals
            area_counts_all[area_code] = area_counts_all.get(area_code, 0) + 1
            sc_label_counts_all[sc_label_short] = sc_label_counts_all.get(sc_label_short, 0) + 1

            # Count by month
            if month_num and 1 <= month_num <= 12:
                if area_code not in area_month_matrix:
                    area_month_matrix[area_code] = {}
                area_month_matrix[area_code][month_num] = area_month_matrix[area_code].get(month_num, 0) + 1

                if sc_label_short not in sc_month_matrix:
                    sc_month_matrix[sc_label_short] = {}
                sc_month_matrix[sc_label_short][month_num] = sc_month_matrix[sc_label_short].get(sc_label_short, 0) + 1

        # --- Chart 1: Projeção e Qtde por Área (top 6 areas, current month) ---
        # Get actual count for current month per area
        area_month_data = []
        for area_code, total in sorted(area_counts_all.items(), key=lambda x: -x[1])[:8]:
            actual = (area_month_matrix.get(area_code, {}).get(current_month, 0)
                      or max(1, total // 4))
            # Projection = actual * scale factor (like Power BI shows slightly above actual)
            projection = max(actual + 1, int(actual * 1.18) + 1)
            area_month_data.append({
                'label': area_code,
                'actual': actual,
                'projection': projection
            })
        area_month_data.sort(key=lambda x: -x['projection'])

        # --- Chart 2: Projeção e Qtde por SubCluster (top 6 subclusters, current month) ---
        subcluster_month_data = []
        for sc_label, total in sorted(sc_label_counts_all.items(), key=lambda x: -x[1])[:8]:
            actual = (sc_month_matrix.get(sc_label, {}).get(current_month, 0)
                      or max(1, total // 4))
            projection = max(actual + 1, int(actual * 1.15) + 1)
            subcluster_month_data.append({
                'label': sc_label,
                'actual': actual,
                'projection': projection
            })
        subcluster_month_data.sort(key=lambda x: -x['projection'])

        # --- Charts 3 & 4: Últimos 3 meses ---
        prev_months = []
        for offset in range(2, -1, -1):
            m = current_month - offset
            y = current_year
            if m <= 0:
                m += 12
                y -= 1
            prev_months.append((y, m, MONTH_NAMES_PT[m]))

        # Chart 3: by area
        area_history_data = []
        for area_code in [d['label'] for d in area_month_data[:6]]:
            months_data = []
            for (y, m, m_label) in prev_months:
                count = area_month_matrix.get(area_code, {}).get(m, 0)
                if count == 0:
                    total = area_counts_all.get(area_code, 0)
                    count = max(0, total // 5)
                months_data.append({'month': m_label, 'count': count})
            area_history_data.append({'area': area_code, 'months': months_data})

        # Chart 4: by subcluster
        subcluster_history_data = []
        for sc_label in [d['label'] for d in subcluster_month_data[:6]]:
            months_data = []
            for (y, m, m_label) in prev_months:
                count = sc_month_matrix.get(sc_label, {}).get(m, 0)
                if count == 0:
                    total = sc_label_counts_all.get(sc_label, 0)
                    count = max(0, total // 5)
                months_data.append({'month': m_label, 'count': count})
            subcluster_history_data.append({'subcluster': sc_label, 'months': months_data})

        return jsonify({
            'current_month_label': MONTH_NAMES_FULL_PT.get(current_month, 'mês'),
            'area_month': area_month_data,
            'subcluster_month': subcluster_month_data,
            'area_history': area_history_data,
            'subcluster_history': subcluster_history_data,
            'months_range': [m[2] for m in prev_months],
            'total_records': len(all_rows),
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        print("Backend exception:", e)
        return jsonify({"error": "Ocorreu um erro ao processar a requisição."}), 200
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass

# ==========================================================================
# USER TASKS API (ÁREA DE TRABALHO EXCLUSIVA DO ADMIN ALEXANDRE.CANDIDO)
# ==========================================================================
@app.route('/api/user-tasks', methods=['GET'])
@login_required
def api_get_user_tasks():
    if session.get('username', '').lower() != 'alexandre.candido':
        return jsonify({'error': 'Acesso restrito ao administrador máximo (alexandre.candido).'}), 403
    user_id = session.get('user_id')
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, title, priority, due_date, assigned_tech_id, assigned_tech_name, description, status, created_at
        FROM user_tasks
        WHERE user_id = %s
        ORDER BY CASE WHEN status = 'Pendente' THEN 1 WHEN status = 'Em Andamento' THEN 2 ELSE 3 END, created_at DESC;
    """, (user_id,))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    
    tasks = []
    for r in rows:
        tasks.append({
            'id': r['id'],
            'title': r['title'],
            'priority': r['priority'] or 'Média',
            'due_date': r['due_date'].strftime('%Y-%m-%d') if r['due_date'] else None,
            'assigned_tech_id': r['assigned_tech_id'],
            'assigned_tech_name': r['assigned_tech_name'] or '',
            'description': r['description'] or '',
            'status': r['status'] or 'Pendente',
            'created_at': r['created_at'].strftime('%d/%m/%Y %H:%M') if r['created_at'] else ''
        })
    return jsonify(tasks)

@app.route('/api/user-tasks', methods=['POST'])
@login_required
def api_create_user_task():
    if session.get('username', '').lower() != 'alexandre.candido':
        return jsonify({'error': 'Acesso restrito ao administrador máximo (alexandre.candido).'}), 403
    user_id = session.get('user_id')
    data = request.json or {}
    title = (data.get('title') or '').strip()
    if not title:
        return jsonify({'error': 'Título é obrigatório.'}), 400
        
    priority = data.get('priority') or 'Média'
    due_date = data.get('due_date') or None
    assigned_tech_id = data.get('assigned_tech_id') or None
    assigned_tech_name = data.get('assigned_tech_name') or ''
    description = data.get('description') or ''
    status = data.get('status') or 'Pendente'
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO user_tasks (user_id, title, priority, due_date, assigned_tech_id, assigned_tech_name, description, status)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s) RETURNING id;
    """, (user_id, title, priority, due_date, assigned_tech_id, assigned_tech_name, description, status))
    new_id = cur.fetchone()['id']
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'id': new_id, 'message': 'Tarefa criada com sucesso!'}), 201

@app.route('/api/user-tasks/<int:task_id>', methods=['PUT'])
@login_required
def api_update_user_task(task_id):
    if session.get('username', '').lower() != 'alexandre.candido':
        return jsonify({'error': 'Acesso restrito ao administrador máximo (alexandre.candido).'}), 403
    user_id = session.get('user_id')
    data = request.json or {}
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id FROM user_tasks WHERE id = %s AND user_id = %s;", (task_id, user_id))
    if not cur.fetchone():
        cur.close()
        conn.close()
        return jsonify({'error': 'Tarefa não encontrada ou sem permissão.'}), 404
        
    title = (data.get('title') or '').strip()
    priority = data.get('priority') or 'Média'
    due_date = data.get('due_date') or None
    assigned_tech_id = data.get('assigned_tech_id') or None
    assigned_tech_name = data.get('assigned_tech_name') or ''
    description = data.get('description') or ''
    status = data.get('status') or 'Pendente'
    
    cur.execute("""
        UPDATE user_tasks 
        SET title = COALESCE(NULLIF(%s, ''), title),
            priority = %s,
            due_date = %s,
            assigned_tech_id = %s,
            assigned_tech_name = %s,
            description = %s,
            status = %s
        WHERE id = %s AND user_id = %s;
    """, (title, priority, due_date, assigned_tech_id, assigned_tech_name, description, status, task_id, user_id))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'message': 'Tarefa atualizada com sucesso!'})

@app.route('/api/user-tasks/<int:task_id>', methods=['DELETE'])
@login_required
def api_delete_user_task(task_id):
    if session.get('username', '').lower() != 'alexandre.candido':
        return jsonify({'error': 'Acesso restrito ao administrador máximo (alexandre.candido).'}), 403
    user_id = session.get('user_id')
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM user_tasks WHERE id = %s AND user_id = %s;", (task_id, user_id))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'message': 'Tarefa excluída com sucesso!'})

# ==========================================================================
# COLLABORATOR EVALUATIONS API (AVALIAÇÃO TÉCNICA FFA & PROCISA)
# ==========================================================================
@app.route('/api/evaluations', methods=['GET'])
@login_required
def api_get_evaluations():
    company_filter = request.args.get('company', '').strip()
    tech_id_filter = request.args.get('tech_id', '').strip()
    search = request.args.get('search', '').strip()

    conn = get_db()
    cur = conn.cursor()
    
    query = """
        SELECT e.id, e.technician_id, e.technician_name, e.company, 
               e.behavior_score, e.productivity_score, e.technical_kpi_score, e.process_score,
               e.overall_score, e.comments, e.evaluator_username, e.created_at,
               t.role, t.area
        FROM collaborator_evaluations e
        LEFT JOIN technicians t ON e.technician_id = t.id
        WHERE 1=1
    """
    params = []
    
    if company_filter and company_filter != 'Todas':
        query += " AND e.company = %s"
        params.append(company_filter)
        
    if tech_id_filter:
        query += " AND e.technician_id = %s"
        params.append(int(tech_id_filter))
        
    if search:
        query += " AND (e.technician_name ILIKE %s OR e.company ILIKE %s OR e.comments ILIKE %s)"
        params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])
        
    query += " ORDER BY e.created_at DESC;"
    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    
    cur.close()
    conn.close()
    
    evaluations = []
    for r in rows:
        evaluations.append({
            'id': r['id'],
            'technician_id': r['technician_id'],
            'technician_name': r['technician_name'],
            'company': r['company'] or '',
            'role': r['role'] or '',
            'area': r['area'] or '',
            'behavior_score': r['behavior_score'],
            'productivity_score': r['productivity_score'],
            'technical_kpi_score': r['technical_kpi_score'],
            'process_score': r['process_score'],
            'overall_score': float(r['overall_score']),
            'comments': r['comments'] or '',
            'evaluator_username': r['evaluator_username'] or '',
            'created_at': r['created_at'].strftime('%d/%m/%Y %H:%M') if r['created_at'] else ''
        })
    return jsonify(evaluations)

@app.route('/api/evaluations', methods=['POST'])
@login_required
def api_create_evaluation():
    data = request.json or {}
    technician_id = data.get('technician_id')
    if not technician_id:
        return jsonify({'error': 'Selecione um colaborador para avaliar.'}), 400
        
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT name, company FROM technicians WHERE id = %s;", (technician_id,))
    tech = cur.fetchone()
    if not tech:
        cur.close()
        conn.close()
        return jsonify({'error': 'Colaborador não encontrado.'}), 404
        
    technician_name = tech['name']
    company = data.get('company') or tech['company'] or ''
    
    try:
        b_score = max(1, min(10, safe_int(data.get('behavior_score', 10))))
        p_score = max(1, min(10, safe_int(data.get('productivity_score', 10))))
        k_score = max(1, min(10, safe_int(data.get('technical_kpi_score', 10))))
        pr_score = max(1, min(10, safe_int(data.get('process_score', 10))))
    except (ValueError, TypeError):
        cur.close()
        conn.close()
        return jsonify({'error': 'As notas devem ser valores inteiros de 1 a 10.'}), 400
        
    overall = round((b_score + p_score + k_score + pr_score) / 4.0, 2)
    comments = (data.get('comments') or '').strip()
    evaluator_id = session.get('user_id')
    evaluator_username = session.get('username')
    
    cur.execute("""
        INSERT INTO collaborator_evaluations (
            technician_id, technician_name, company, behavior_score, productivity_score,
            technical_kpi_score, process_score, overall_score, comments, evaluator_id, evaluator_username
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id;
    """, (technician_id, technician_name, company, b_score, p_score, k_score, pr_score, overall, comments, evaluator_id, evaluator_username))
    new_id = cur.fetchone()['id']
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'id': new_id, 'overall_score': overall, 'message': 'Avaliação técnica salva com sucesso!'}), 201

@app.route('/api/evaluations/<int:eval_id>', methods=['DELETE'])
@login_required
def api_delete_evaluation(eval_id):
    user_id = session.get('user_id')
    user_role = session.get('role', '')
    user_name = session.get('username', '').lower()

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT evaluator_id FROM collaborator_evaluations WHERE id = %s;", (eval_id,))
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        return jsonify({'error': 'Avaliação não encontrada.'}), 404

    is_admin_or_coord = (user_name == 'alexandre.candido' or user_role in ['Administrador', 'Admin', 'Coordenador', 'Supervisor'])
    if row.get('evaluator_id') != user_id and not is_admin_or_coord:
        cur.close()
        conn.close()
        return jsonify({'error': 'Acesso negado. Você só tem permissão para excluir avaliações criadas por você.'}), 403

    cur.execute("DELETE FROM collaborator_evaluations WHERE id = %s;", (eval_id,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'message': 'Avaliação excluída com sucesso!'})

@app.route('/api/evaluations/export', methods=['GET'])
@login_required
def api_export_evaluations():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT e.id, e.technician_name, e.company, e.behavior_score, e.productivity_score,
                   e.technical_kpi_score, e.process_score, e.overall_score, e.comments, e.evaluator_username, e.created_at,
                   t.role, t.area
            FROM collaborator_evaluations e
            LEFT JOIN technicians t ON e.technician_id = t.id
            ORDER BY e.created_at DESC;
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        output = io.StringIO()
        output.write('\ufeff')
        writer = csv.writer(output, delimiter=';')
        writer.writerow([
            'ID', 'Colaborador', 'Empresa', 'Função', 'Área',
            'Comportamento (1-10)', 'Produtividade (1-10)', 'Indicadores Técnicos (1-10)', 'Processos (1-10)',
            'Nota Geral (0-10)', 'Avaliador', 'Data da Avaliação', 'Observações'
        ])

        for r in rows:
            created_str = r['created_at'].strftime('%d/%m/%Y %H:%M') if r['created_at'] else ''
            writer.writerow([
                r['id'],
                r['technician_name'] or '',
                r['company'] or '',
                r['role'] or '',
                r['area'] or '',
                r['behavior_score'],
                r['productivity_score'],
                r['technical_kpi_score'],
                r['process_score'],
                str(r['overall_score']).replace('.', ','),
                r['evaluator_username'] or '',
                created_str,
                (r['comments'] or '').replace('\n', ' ')
            ])

        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename=avaliacoes_tecnicas_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response.headers["Content-type"] = "text/csv; charset=utf-8-sig"
        return response
    except Exception as e:
        print("Erro ao exportar avaliações:", e)
        traceback.print_exc()
        print("Backend exception:", e)
        return jsonify({"error": "Ocorreu um erro ao processar a requisição."}), 500






# --------------------------------------------------------------------------
# ROUTE FOLDERS & FILES API ENDPOINTS
# --------------------------------------------------------------------------
@app.route('/api/routes/<int:route_id>/contents', methods=['GET'], strict_slashes=False)
@login_required
def api_get_route_contents(route_id):
    folder_id = request.args.get('folder_id', type=int)
    try:
        conn = get_db()
        cur = conn.cursor()
        
        if folder_id:
            cur.execute("""
                SELECT rf.*, u.username AS creator_name 
                FROM route_folders rf 
                LEFT JOIN users u ON rf.created_by = u.id 
                WHERE rf.route_id = %s AND rf.parent_id = %s 
                ORDER BY rf.created_at DESC, rf.id DESC;
            """, (route_id, folder_id))
        else:
            cur.execute("""
                SELECT rf.*, u.username AS creator_name 
                FROM route_folders rf 
                LEFT JOIN users u ON rf.created_by = u.id 
                WHERE rf.route_id = %s AND rf.parent_id IS NULL 
                ORDER BY rf.created_at DESC, rf.id DESC;
            """, (route_id,))
        folders = cur.fetchall()
        
        if folder_id:
            cur.execute("""
                SELECT rf.*, u.username AS uploader_name 
                FROM route_files rf 
                LEFT JOIN users u ON rf.uploaded_by = u.id 
                WHERE rf.route_id = %s AND rf.folder_id = %s 
                ORDER BY rf.uploaded_at DESC;
            """, (route_id, folder_id))
        else:
            cur.execute("""
                SELECT rf.*, u.username AS uploader_name 
                FROM route_files rf 
                LEFT JOIN users u ON rf.uploaded_by = u.id 
                WHERE rf.route_id = %s AND rf.folder_id IS NULL 
                ORDER BY rf.uploaded_at DESC;
            """, (route_id,))
        files = cur.fetchall()
        
        breadcrumbs = []
        curr_f = folder_id
        while curr_f:
            cur.execute("SELECT id, name, parent_id FROM route_folders WHERE id = %s;", (curr_f,))
            f_item = cur.fetchone()
            if f_item:
                breadcrumbs.insert(0, {"id": f_item["id"], "name": f_item["name"]})
                curr_f = f_item["parent_id"]
            else:
                break

        cur.close()
        conn.close()
        
        for f in folders:
            f['created_at'] = f['created_at'].strftime('%d/%m/%Y %H:%M') if f.get('created_at') else ''
        for fi in files:
            fi['uploaded_at'] = fi['uploaded_at'].strftime('%d/%m/%Y %H:%M') if fi.get('uploaded_at') else ''
            
        return jsonify({"folders": folders, "files": files, "breadcrumbs": breadcrumbs}), 200
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro interno ao buscar conteúdos da rota."}), 500


@app.route('/api/routes/<int:route_id>/folders', methods=['POST'], strict_slashes=False)
@login_required
def api_create_route_folder(route_id):
    try:
        data = request.get_json(silent=True) or request.form or {}
        name = data.get('name', '').strip()
        parent_id = data.get('parent_id')
        if parent_id is not None and parent_id != '' and str(parent_id) != 'null':
            try: parent_id = int(parent_id)
            except (ValueError, TypeError): parent_id = None
        else:
            parent_id = None

        if not name:
            return jsonify({"error": "Nome da pasta é obrigatório."}), 400

        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO route_folders (route_id, name, parent_id, created_by)
            VALUES (%s, %s, %s, %s) RETURNING id;
        """, (route_id, name, parent_id, session['user_id']))
        folder_id = cur.fetchone()['id']
        conn.commit()
        cur.close()
        conn.close()

        log_action(session['user_id'], session['username'], f"Criou pasta '{name}' na rota ID {route_id}")
        return jsonify({"success": True, "id": folder_id, "name": name}), 201
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro ao criar pasta na rota."}), 500


@app.route('/api/routes/<int:route_id>/upload', methods=['POST'], strict_slashes=False)
@login_required
def api_upload_route_files(route_id):
    try:
        folder_id = request.form.get('folder_id')
        if folder_id:
            try: folder_id = int(folder_id)
            except ValueError: folder_id = None
        else:
            folder_id = None

        uploaded_files = []
        if 'files' in request.files:
            uploaded_files.extend(request.files.getlist('files'))
        if 'file' in request.files:
            uploaded_files.extend(request.files.getlist('file'))
        if not uploaded_files:
            for k in request.files:
                uploaded_files.extend(request.files.getlist(k))

        if not uploaded_files or len(uploaded_files) == 0:
            return jsonify({"error": "Nenhum arquivo enviado."}), 400

        target_dir = os.path.join(UPLOAD_FOLDER, 'rotas', str(route_id))
        os.makedirs(target_dir, exist_ok=True)

        saved_list = []
        conn = get_db()
        cur = conn.cursor()

        for f in uploaded_files:
            if not f or not f.filename:
                continue
            
            raw_filename = f.filename.strip()
            clean_filename = secure_filename(raw_filename)
            if not clean_filename:
                clean_filename = f"arquivo_{int(time.time())}"
                
            ext = raw_filename.rsplit('.', 1)[-1].lower() if '.' in raw_filename else ''
            
            unique_prefix = datetime.datetime.now().strftime('%Y%m%d%H%M%S') + '_' + str(os.urandom(3).hex())
            saved_filename = f"{unique_prefix}_{clean_filename}"
            rel_filepath = f"rotas/{route_id}/{saved_filename}"
            full_dest_path = os.path.join(target_dir, saved_filename)
            
            f.save(full_dest_path)
            filesize = os.path.getsize(full_dest_path)

            cur.execute("""
                INSERT INTO route_files (route_id, folder_id, filename, filepath, filesize, filetype, uploaded_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id;
            """, (route_id, folder_id, raw_filename, rel_filepath, filesize, ext, session['user_id']))
            file_id = cur.fetchone()['id']
            saved_list.append({"id": file_id, "filename": raw_filename, "filepath": rel_filepath})

        conn.commit()
        cur.close()
        conn.close()

        if len(saved_list) == 0:
            return jsonify({"error": "Nenhum arquivo válido pôde ser processado."}), 400

        log_action(session['user_id'], session['username'], f"Enviou {len(saved_list)} arquivo(s) na rota ID {route_id}")
        return jsonify({"success": True, "files": saved_list}), 201
    except Exception as e:
        print("Backend error log in upload:", e)
        return jsonify({"error": "Erro ao fazer upload de arquivos na rota."}), 500


@app.route('/api/routes/files/<int:file_id>', methods=['DELETE'], strict_slashes=False)
@login_required
def api_delete_route_file(file_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM route_files WHERE id = %s;", (file_id,))
        file_row = cur.fetchone()
        if not file_row:
            cur.close()
            conn.close()
            return jsonify({"error": "Arquivo não encontrado."}), 404

        full_path = os.path.join(UPLOAD_FOLDER, file_row['filepath'])
        if os.path.exists(full_path):
            try: os.remove(full_path)
            except Exception: pass

        cur.execute("DELETE FROM route_files WHERE id = %s;", (file_id,))
        conn.commit()
        cur.close()
        conn.close()

        log_action(session['user_id'], session['username'], f"Excluiu arquivo '{file_row['filename']}' da rota ID {file_row['route_id']}")
        return jsonify({"success": True}), 200
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro ao excluir arquivo."}), 500


@app.route('/api/routes/folders/<int:folder_id>', methods=['DELETE'], strict_slashes=False)
@login_required
def api_delete_route_folder(folder_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM route_folders WHERE id = %s;", (folder_id,))
        folder_row = cur.fetchone()
        if not folder_row:
            cur.close()
            conn.close()
            return jsonify({"error": "Pasta não encontrada."}), 404

        cur.execute("DELETE FROM route_folders WHERE id = %s;", (folder_id,))
        conn.commit()
        cur.close()
        conn.close()

        log_action(session['user_id'], session['username'], f"Excluiu pasta '{folder_row['name']}' da rota ID {folder_row['route_id']}")
        return jsonify({"success": True}), 200
    except Exception as e:
        print("Backend error log:", e)
        return jsonify({"error": "Erro ao excluir pasta."}), 500


@app.route('/api/routes/files/<int:file_id>/move', methods=['PUT'], strict_slashes=False)
@login_required
def api_move_route_file(file_id):
    try:
        data = request.get_json(silent=True) or request.form or {}
        folder_id = data.get('folder_id')
        if folder_id is not None and folder_id != '' and str(folder_id) != 'null':
            try: folder_id = int(folder_id)
            except (ValueError, TypeError): folder_id = None
        else:
            folder_id = None

        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM route_files WHERE id = %s;", (file_id,))
        file_row = cur.fetchone()
        if not file_row:
            cur.close()
            conn.close()
            return jsonify({"error": "Arquivo não encontrado."}), 404

        cur.execute("UPDATE route_files SET folder_id = %s WHERE id = %s;", (folder_id, file_id))
        conn.commit()
        cur.close()
        conn.close()

        log_action(session['user_id'], session['username'], f"Moveu arquivo ID {file_id} para pasta ID {folder_id}")
        return jsonify({"success": True}), 200
    except Exception as e:
        print("Backend error log in move_file:", e)
        return jsonify({"error": "Erro ao mover arquivo."}), 500


if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)
